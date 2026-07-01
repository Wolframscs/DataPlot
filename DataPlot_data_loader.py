import os
import time
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl

try:
    from python_calamine import CalamineWorkbook
    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False

class DataLoaderMixin:
    def process_data(self):
        """处理Excel/CSV数据"""
        try:
            file_path = self.file_path.get()
            if not os.path.exists(file_path):
                self.update_status("错误：文件不存在", clear=True)
                return
                
            self.set_buttons_state(False)
            
            file_type = self.file_type.get()
            sheet_name = self.sheet_name.get()
            start_row_str = self.start_row.get()
            
            threading.Thread(
                target=self._bg_process_data,
                args=(file_path, file_type, sheet_name, start_row_str),
                daemon=True
            ).start()
            
        except Exception as e:
            error_msg = f"启动处理线程失败: {str(e)}"
            self.logger.error(error_msg)
            self.update_status(error_msg, clear=True)
            self.set_buttons_state(True)

    def _bg_process_data(self, file_path, file_type, sheet_name, start_row_str):
        """后台线程处理函数，避免阻塞 UI 主线程"""
        try:
            try:
                skip_rows = int(self.skip_rows_var.get())
            except ValueError:
                skip_rows = 0
                
            try:
                start_skip = int(self.start_skip_var.get())
            except ValueError:
                start_skip = 0

            if file_type == "raw":
                self.msg_queue.put({'type': 'status', 'message': "开始读取原始Excel文件...", 'clear': True})
                try:
                    start_time = time.time()
                    self.msg_queue.put({'type': 'status', 'message': "正在读取文件..."})
                    df = None

                    try:
                        self.msg_queue.put({'type': 'status', 'message': "优先尝试使用 Polars 快速读取 Excel..."})
                        import polars as pl
                        df_pl = pl.read_excel(
                            file_path,
                            sheet_name=sheet_name,
                            read_options={'header_row': skip_rows}
                        )
                        df = pd.DataFrame(df_pl.to_dict(as_series=False))
                        df.columns = [col.replace('__UNNAMED__', 'Unnamed: ') if str(col).startswith('__UNNAMED__') else col for col in df.columns]
                    except Exception as pl_err:
                        self.msg_queue.put({'type': 'status', 'message': f"Polars 读取 Excel 失败: {str(pl_err)}，切换至 Pandas Calamine 读取..."})
                        df = None
                        
                    if df is None:
                        engine = 'calamine' if HAS_CALAMINE else 'openpyxl'
                        df = pd.read_excel(
                            file_path,
                            sheet_name=sheet_name,
                            skiprows=skip_rows,
                            header=0,
                            engine=engine
                        )
                    
                    df = df.iloc[start_skip:].reset_index(drop=True)
                    
                    def clean_column_name(col):
                        col = str(col)
                        if '（' not in col and '(' not in col:
                            return col
                        
                        parts = col.split()
                        if len(parts) > 1 and parts[-1].isdigit():
                            number = parts[-1]
                            prefix = ' '.join(parts[:-1])
                            if '（' in prefix or '(' in prefix:
                                pos1 = prefix.find('（')
                                pos2 = prefix.find('(')
                                if pos1 == -1: pos1 = len(prefix)
                                if pos2 == -1: pos2 = len(prefix)
                                pos = min(pos1, pos2)
                                prefix = prefix[:pos]
                            return f"{prefix.strip()} {number}"
                        else:
                            pos1 = col.find('（')
                            pos2 = col.find('(')
                            if pos1 == -1: pos1 = len(col)
                            if pos2 == -1: pos2 = len(col)
                            pos = min(pos1, pos2)
                            return col[:pos].strip()
                    
                    df.columns = [clean_column_name(col) for col in df.columns]
                    
                    result_df = pd.DataFrame()
                    result_df['Time'] = df.iloc[:, 0]
                    
                    # Only read the first column as Time, and subsequent non-time, non-empty columns as data
                    for col_idx in range(1, len(df.columns)):
                        col = df.columns[col_idx]
                        col_str = str(col).lower().strip()
                        if 'unnamed' in col_str or col_str == '':
                            continue
                        if any(x in col_str for x in ['时间', 'time', 'ʱ', '[s]', '(s)']):
                            continue
                        result_df[col] = df.iloc[:, col_idx]
                    
                    # Convert all columns to numeric types
                    for col in result_df.columns:
                        result_df[col] = pd.to_numeric(result_df[col], errors='coerce')
                    
                    end_time = time.time()
                    elapsed_time = end_time - start_time
                    self.msg_queue.put({'type': 'status', 'message': f"文件读取完成，耗时: {elapsed_time:.2f}秒"})
                    
                    self.msg_queue.put({'type': 'status', 'message': "数据读取与清洗完成"})
                    self.msg_queue.put({'type': 'raw_done', 'df': result_df, 'message': "数据读取与清洗完成"})
                    
                except Exception as e:
                    self.msg_queue.put({'type': 'error', 'message': f"读取Excel文件失败: {str(e)}"})
                    return

            elif file_type == "battery":
                self.msg_queue.put({'type': 'status', 'message': "开始读取电池数据...", 'clear': True})
                try:
                    start_time = time.time()
                    file_ext = os.path.splitext(file_path)[1].lower()
                    
                    try:
                        start_row = int(start_row_str) - 1
                        if start_row < 0:
                            start_row = 0
                    except ValueError:
                        start_row = 0
                    
                    start_row = start_row + skip_rows

                    if file_ext == '.csv':
                        self.msg_queue.put({'type': 'status', 'message': "正在读取CSV文件..."})
                        df = None
                        
                        try:
                            self.msg_queue.put({'type': 'status', 'message': "尝试使用 Polars 极速读取..."})
                            import polars as pl
                            df_pl = pl.read_csv(file_path, skip_rows=start_row, ignore_errors=True, infer_schema_length=10000)
                            df = pd.DataFrame(df_pl.to_dict(as_series=False))
                            if not df.empty:
                                end_time = time.time()
                                self.msg_queue.put({'type': 'status', 'message': f"使用 Polars 成功读取文件，耗时: {end_time - start_time:.2f}秒"})
                        except ImportError:
                            self.msg_queue.put({'type': 'status', 'message': "未安装 Polars，切换至 Pandas 兼容模式..."})
                        except Exception as pl_err:
                            self.msg_queue.put({'type': 'status', 'message': f"Polars 读取失败（可能编码非 UTF-8 或格式不符）: {str(pl_err)}，切换至 Pandas 兼容模式..."})
                            df = None
                        
                        if df is None:
                            try:
                                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                    total_lines = sum(1 for _ in f)
                                self.msg_queue.put({'type': 'status', 'message': f"文件总行数: {total_lines}"})
                            except Exception:
                                total_lines = None
                            
                            encodings = ['utf-8', 'gb18030', 'gbk', 'gb2312', 'ascii']
                            
                            for encoding in encodings:
                                try:
                                    self.msg_queue.put({'type': 'status', 'message': f"尝试使用 {encoding} 编码读取..."})
                                    chunks = []
                                    for i, chunk in enumerate(pd.read_csv(file_path, encoding=encoding, 
                                                                         chunksize=self.CHUNK_SIZE, 
                                                                         skiprows=start_row if start_row > 0 else None)):
                                        chunks.append(chunk)
                                        if total_lines:
                                            current_rows = min((i + 1) * self.CHUNK_SIZE, total_lines)
                                            progress = min(100, current_rows / total_lines * 100)
                                            self.msg_queue.put({'type': 'status', 'message': f"读取进度: {progress:.1f}% ({current_rows}/{total_lines}行)"})
                                    
                                    df = pd.concat(chunks, ignore_index=True)
                                    if not df.empty:
                                        end_time = time.time()
                                        self.msg_queue.put({'type': 'status', 'message': f"使用 {encoding} 编码成功读取文件，耗时: {end_time - start_time:.2f}秒"})
                                        break
                                except Exception as e:
                                    self.msg_queue.put({'type': 'status', 'message': f"使用 {encoding} 编码读取失败: {str(e)}"})
                                    continue
                                    
                        if df is None:
                            self.msg_queue.put({'type': 'error', 'message': "读取电池CSV文件失败，所有编码尝试均已失败。"})
                            return
                        df = df.dropna(how='all')
                    else:
                        self.msg_queue.put({'type': 'status', 'message': "正在读取文件..."})
                        df = None
                        try:
                            self.msg_queue.put({'type': 'status', 'message': "优先尝试使用 Polars 快速读取 Excel..."})
                            import polars as pl
                            read_opts = {'header_row': start_row} if start_row > 0 else {}
                            df_pl = pl.read_excel(
                                file_path,
                                sheet_name=sheet_name,
                                read_options=read_opts
                            )
                            df = pd.DataFrame(df_pl.to_dict(as_series=False))
                            df.columns = [col.replace('__UNNAMED__', 'Unnamed: ') if str(col).startswith('__UNNAMED__') else col for col in df.columns]
                            df = df.dropna(how='all')
                        except Exception as pl_err:
                            self.msg_queue.put({'type': 'status', 'message': f"Polars 读取 Excel 失败: {str(pl_err)}，切换至 Pandas Calamine 读取..."})
                            df = None

                        if df is None:
                            use_calamine = HAS_CALAMINE
                            if use_calamine:
                                self.msg_queue.put({'type': 'status', 'message': "正在使用 Calamine 快速读取..."})
                                try:
                                    df = pd.read_excel(
                                        file_path,
                                        sheet_name=sheet_name,
                                        skiprows=start_row if start_row > 0 else None,
                                        engine='calamine'
                                    )
                                    df = df.dropna(how='all')
                                except Exception as e:
                                    self.msg_queue.put({'type': 'status', 'message': f"Calamine 读取失败, 尝试标准方法: {str(e)}"})
                                    use_calamine = False
                            
                            if not use_calamine:
                                self.msg_queue.put({'type': 'status', 'message': "正在使用 openpyxl read-only 模式读取..."})
                                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                                sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                                total_rows = sheet.max_row
                                self.msg_queue.put({'type': 'status', 'message': f"文件总行数: {total_rows or '未知'}"})
                                
                                data = []
                                header = None
                                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                                    if r_idx < start_row:
                                        continue
                                    if r_idx == start_row:
                                        header = list(row)
                                    else:
                                        data.append(row)
                                    
                                    if r_idx > start_row and (r_idx - start_row) % 20000 == 0:
                                        percent = ((r_idx - start_row) / total_rows * 100) if total_rows else 0
                                        self.msg_queue.put({'type': 'status', 'message': f"读取进度: {percent:.1f}% ({r_idx - start_row}/{total_rows - start_row if total_rows else '未知'}行)"})
                                
                                df = pd.DataFrame(data, columns=header)
                                df = df.dropna(how='all')
                                wb.close()

                    if df is not None and start_skip > 0:
                        df = df.iloc[start_skip:].reset_index(drop=True)

                    cycle_default = None
                    step_default = None
                    time_default = None
                    voltage_default = None
                    current_default = None
                    
                    for col in df.columns:
                        col_lower = str(col).lower()
                        if cycle_default is None and any(x in col_lower for x in ['循环序号', '循环号', '循环', 'cycle']):
                            cycle_default = col
                        if step_default is None and any(x in col_lower for x in ['工步状态', '工步', '工作状态', 'step', 'state', 'status']):
                            step_default = col
                        if time_default is None and any(x in col_lower for x in ['时间', 'time']):
                            if not col_lower.endswith('_时间差(s)'):
                                time_default = col
                        if voltage_default is None and (any(x in col_lower for x in ['电压', 'voltage', 'volt', 'u(v)']) or col_lower in ['u', 'v']):
                            voltage_default = col
                        if current_default is None and (any(x in col_lower for x in ['电流', 'current', 'curr', 'i(a)', 'i(ma)']) or col_lower in ['i']):
                            current_default = col

                    if not cycle_default and len(df.columns) > 0:
                        cycle_default = df.columns[0]
                    if not step_default and len(df.columns) > 0:
                        step_default = df.columns[min(1, len(df.columns)-1)]
                    if not time_default and len(df.columns) > 0:
                        time_default = df.columns[min(2, len(df.columns)-1)]
                    if not voltage_default and len(df.columns) > 0:
                        voltage_default = df.columns[min(3, len(df.columns)-1)]
                    if not current_default and len(df.columns) > 0:
                        current_default = df.columns[min(4, len(df.columns)-1)]

                    import pandas.api.types as ptypes
                    import warnings
                    self.msg_queue.put({'type': 'status', 'message': "正在转换默认时间列并计算时间差(s)..."})
                    
                    if time_default and time_default in df.columns:
                        col = time_default
                        new_col_name = f"{col}_时间差(s)"
                        if ptypes.is_datetime64_any_dtype(df[col]):
                            first_valid = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
                            if first_valid is not None:
                                df[new_col_name] = (df[col] - first_valid).dt.total_seconds()
                        elif df[col].dtype == 'object' or ptypes.is_string_dtype(df[col]):
                            non_nulls = df[col].dropna()
                            if not non_nulls.empty:
                                try:
                                    with warnings.catch_warnings():
                                        warnings.filterwarnings("ignore", category=UserWarning, message=".*Could not infer format.*")
                                        pd.to_datetime(non_nulls.head(10), errors='raise')
                                        times = pd.to_datetime(df[col], errors='coerce')
                                    first_valid = times.dropna().iloc[0] if not times.dropna().empty else None
                                    if first_valid is not None:
                                        df[new_col_name] = (times - first_valid).dt.total_seconds()
                                except Exception:
                                    pass

                    self.msg_queue.put({'type': 'status', 'message': "正在生成工步 and 循环筛选项..."})
                    cycles = ['全部']
                    if cycle_default:
                        try:
                            unique_vals = df[cycle_default].dropna().unique()
                            def safe_sort_key(val):
                                try:
                                    return (0, float(val))
                                except (ValueError, TypeError):
                                    return (1, str(val))
                            unique_vals = sorted(unique_vals, key=safe_sort_key)
                            for x in unique_vals:
                                try:
                                    if float(x).is_integer():
                                        cycles.append(str(int(x)))
                                    else:
                                        cycles.append(str(x))
                                except ValueError:
                                    cycles.append(str(x))
                        except Exception:
                            cycles = ['全部'] + [str(x) for x in df[cycle_default].dropna().unique()]
                            
                    steps = ['全部']
                    if step_default:
                        try:
                            steps = ['全部'] + [str(x) for x in df[step_default].dropna().unique()]
                        except Exception:
                            pass

                    end_time = time.time()
                    self.msg_queue.put({
                        'type': 'done', 
                        'df': df,
                        'cycles': cycles,
                        'steps': steps,
                        'cycle_default': cycle_default,
                        'step_default': step_default,
                        'time_default': time_default,
                        'voltage_default': voltage_default,
                        'current_default': current_default,
                        'message': f"文件读取完成，耗时: {end_time - start_time:.2f}秒，共读取 {len(df)} 行数据"
                    })
                    
                except Exception as e:
                    self.msg_queue.put({'type': 'error', 'message': f"读取电池数据失败: {str(e)}"})
                    return
            
            else:
                try:
                    start_row = int(start_row_str) - 1
                    if start_row < 0:
                        start_row = 0
                except ValueError:
                    start_row = 0
                
                start_row = start_row + skip_rows
                
                file_ext = os.path.splitext(file_path)[1].lower()
                if file_ext == '.csv':
                    self.msg_queue.put({'type': 'status', 'message': "开始读取CSV文件...", 'clear': True})
                    start_time = time.time()
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            total_lines = sum(1 for _ in f)
                        self.msg_queue.put({'type': 'status', 'message': f"文件总行数: {total_lines}"})
                    except Exception:
                        total_lines = None
                    
                    encodings = ['utf-8', 'gb18030', 'gbk', 'gb2312', 'ascii']
                    result_df = None
                    
                    for encoding in encodings:
                        try:
                            self.msg_queue.put({'type': 'status', 'message': f"尝试使用 {encoding} 编码读取..."})
                            chunks = []
                            for i, chunk in enumerate(pd.read_csv(file_path, encoding=encoding, 
                                                                chunksize=self.CHUNK_SIZE, 
                                                                skiprows=start_row if start_row > 0 else None)):
                                chunks.append(chunk)
                                if total_lines:
                                    current_rows = min((i + 1) * self.CHUNK_SIZE, total_lines)
                                    progress = min(100, current_rows / total_lines * 100)
                                    self.msg_queue.put({'type': 'status', 'message': f"读取进度: {progress:.1f}% ({current_rows}/{total_lines}行)"})
                            
                            result_df = pd.concat(chunks, ignore_index=True)
                            if not result_df.empty:
                                end_time = time.time()
                                self.msg_queue.put({'type': 'status', 'message': f"使用 {encoding} 编码成功读取文件，耗时: {end_time - start_time:.2f}秒"})
                                break
                        except Exception as e:
                            self.msg_queue.put({'type': 'status', 'message': f"使用 {encoding} 编码读取失败: {str(e)}"})
                            continue
                    
                    if result_df is not None:
                        if start_skip > 0:
                            result_df = result_df.iloc[start_skip:].reset_index(drop=True)
                        self.msg_queue.put({'type': 'done', 'df': result_df, 'message': f"CSV文件读取完成，共读取 {len(result_df)} 行数据"})
                    else:
                        self.msg_queue.put({'type': 'error', 'message': "读取CSV文件失败，所有编码尝试均已失败。"})
                            
                elif file_ext in ['.xlsx', '.xls']:
                    self.msg_queue.put({'type': 'status', 'message': "开始读取Excel文件...", 'clear': True})
                    try:
                        start_time = time.time()
                        self.msg_queue.put({'type': 'status', 'message': "正在计算文件大小..."})
                        try:
                            wb = openpyxl.load_workbook(file_path, read_only=True)
                            if sheet_name in wb.sheetnames:
                                sheet = wb[sheet_name]
                                total_rows = sheet.max_row
                            else:
                                total_rows = wb.active.max_row
                            wb.close()
                        except Exception:
                            total_rows = None
                        
                        if total_rows:
                            self.msg_queue.put({'type': 'status', 'message': f"文件总行数: {total_rows}"})
                            if total_rows <= start_row:
                                self.msg_queue.put({'type': 'error', 'message': "错误：起始行超出文件总行数"})
                                return
                        
                        result_df = None
                        try:
                            self.msg_queue.put({'type': 'status', 'message': "优先尝试使用 Polars 快速读取 Excel..."})
                            import polars as pl
                            read_opts = {'header_row': start_row} if start_row > 0 else {}
                            df_pl = pl.read_excel(
                                file_path,
                                sheet_name=sheet_name,
                                read_options=read_opts
                            )
                            result_df = pd.DataFrame(df_pl.to_dict(as_series=False))
                            result_df.columns = [col.replace('__UNNAMED__', 'Unnamed: ') if str(col).startswith('__UNNAMED__') else col for col in result_df.columns]
                        except Exception as pl_err:
                            self.msg_queue.put({'type': 'status', 'message': f"Polars 读取 Excel 失败: {str(pl_err)}，切换至 Pandas Calamine 读取..."})
                            result_df = None

                        if result_df is None:
                            use_calamine = HAS_CALAMINE
                            if use_calamine:
                                self.msg_queue.put({'type': 'status', 'message': "正在使用 Calamine 快速读取..."})
                                try:
                                    result_df = pd.read_excel(
                                        file_path,
                                        sheet_name=sheet_name,
                                        skiprows=start_row,
                                        engine='calamine'
                                    )
                                except Exception as e:
                                    self.msg_queue.put({'type': 'status', 'message': f"Calamine 读取失败, 尝试标准方法: {str(e)}"})
                                    use_calamine = False
                                    
                            if not use_calamine:
                                self.msg_queue.put({'type': 'status', 'message': "正在使用 openpyxl read-only 模式读取..."})
                                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                                sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                                
                                data = []
                                header = None
                                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                                    if r_idx < start_row:
                                        continue
                                    if r_idx == start_row:
                                        header = list(row)
                                    else:
                                        data.append(row)
                                    
                                    if r_idx > start_row and (r_idx - start_row) % 20000 == 0:
                                        percent = ((r_idx - start_row) / total_rows * 100) if total_rows else 0
                                        self.msg_queue.put({'type': 'status', 'message': f"读取进度: {percent:.1f}% ({r_idx - start_row}/{total_rows - start_row if total_rows else '未知'}行)"})
                                
                                result_df = pd.DataFrame(data, columns=header)
                                wb.close()
                        end_time = time.time()
                        if result_df is not None:
                            if start_skip > 0:
                                result_df = result_df.iloc[start_skip:].reset_index(drop=True)
                            self.msg_queue.put({'type': 'done', 'df': result_df, 'message': f"文件读取完成，耗时: {end_time - start_time:.2f}秒，共读取 {len(result_df)} 行数据"})
                        
                    except Exception as e:
                        self.msg_queue.put({'type': 'error', 'message': f"读取Excel文件失败: {str(e)}"})
                        return
                else:
                    self.msg_queue.put({'type': 'error', 'message': "错误：不支持的文件格式"})
                    return
        except Exception as e:
            self.msg_queue.put({'type': 'error', 'message': f"后台处理发生未知错误: {str(e)}"})

    def _bg_calc_time_diff(self, time_col_name, new_col_name):
        try:
            import pandas.api.types as ptypes
            col_data = self.result_df[time_col_name]
            
            numeric_col = pd.to_numeric(col_data, errors='coerce')
            non_null_count = col_data.dropna().shape[0]
            
            if non_null_count > 0 and numeric_col.notna().sum() >= 0.95 * non_null_count:
                non_nulls = numeric_col.dropna()
                first_valid = non_nulls.iloc[0] if not non_nulls.empty else None
                if first_valid is not None:
                    self.result_df[new_col_name] = numeric_col - first_valid
            elif ptypes.is_datetime64_any_dtype(col_data):
                non_nulls = col_data.dropna()
                first_valid = non_nulls.iloc[0] if not non_nulls.empty else None
                if first_valid is not None:
                    self.result_df[new_col_name] = (col_data - first_valid).dt.total_seconds()
            else:
                non_nulls = col_data.dropna()
                if not non_nulls.empty:
                    try:
                        pd.to_datetime(non_nulls.head(10), errors='raise')
                        times = pd.to_datetime(col_data, errors='coerce')
                        valid_times = times.dropna()
                        first_valid = valid_times.iloc[0] if not valid_times.empty else None
                        if first_valid is not None:
                            self.result_df[new_col_name] = (times - first_valid).dt.total_seconds()
                    except Exception:
                        pass
                        
            self.msg_queue.put({
                'type': 'time_diff_done',
                'message': f"计算 {time_col_name} 时间差(s)完成"
            })
        except Exception as e:
            self.msg_queue.put({
                'type': 'error',
                'message': f"计算时间差失败: {str(e)}"
            })

    def browse_file(self):
        if self.file_type.get() == "raw":
            filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]
        else:
            filetypes = [("Excel/CSV files", "*.xlsx;*.csv"), ("Excel files", "*.xlsx"), 
                        ("CSV files", "*.csv"), ("All files", "*.*")]
            
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            self.file_path.set(filename)
            try:
                if filename.endswith('.xlsx'):
                    engine = 'calamine' if HAS_CALAMINE else None
                    excel_file = pd.ExcelFile(filename, engine=engine)
                    self.sheet_combo['values'] = excel_file.sheet_names
                    if excel_file.sheet_names:
                        default_sheet = excel_file.sheet_names[0]
                        for s_name in excel_file.sheet_names:
                            if 'data' in s_name.lower() or '数据' in s_name:
                                default_sheet = s_name
                                break
                        self.sheet_combo.set(default_sheet)
                    self.sheet_combo.configure(state='readonly')
                else:
                    self.sheet_combo.configure(state='disabled')
            except Exception as e:
                messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")

    def update_y_options(self):
        """更新Y轴选项"""
        columns = self.result_df.columns.tolist()
        self.x_combo['values'] = columns
        self.x_combo.set(columns[0])
        
        for listbox in self.y_listboxes:
            listbox.delete(0, tk.END)
            for col in columns[1:]:
                listbox.insert(tk.END, col)

    def convert_csv_to_xlsx(self):
        """将选中的 CSV 文件转换为 XLSX 文件"""
        file_path = self.file_path.get()
        if not file_path:
            messagebox.showwarning("警告", "请先选择需要转换的 CSV 文件！")
            return
            
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext != '.csv':
            messagebox.showwarning("警告", "当前选择的文件不是 CSV 格式！")
            return
            
        self.set_buttons_state(False)
        threading.Thread(target=self._bg_convert_csv_to_xlsx, args=(file_path,), daemon=True).start()

    def _bg_convert_csv_to_xlsx(self, file_path):
        import polars as pl
        import xlsxwriter
        
        filename = os.path.basename(file_path)
        self.msg_queue.put({'type': 'status', 'message': f"开始转换 CSV: {filename}...", 'clear': True})
        
        try:
            start_time = time.time()
            excel_file = os.path.splitext(file_path)[0] + ".xlsx"
            
            MAX_EXCEL_ROWS = 1048575
            workbook = xlsxwriter.Workbook(excel_file, {'constant_memory': True, 'strings_to_numbers': True})
            
            reader = pl.read_csv_batched(file_path, batch_size=MAX_EXCEL_ROWS, ignore_errors=True, infer_schema_length=10000)
            sheet_idx = 1
            total_rows_written = 0
            current_sheet_rows = 0
            worksheet = None
            headers = None
            
            while True:
                batches = reader.next_batches(1)
                if not batches:
                    break
                df = batches[0]
                
                if headers is None:
                    headers = df.columns
                    worksheet = workbook.add_worksheet(f"Sheet{sheet_idx}")
                    worksheet.write_row(0, 0, headers)
                
                for row in df.iter_rows():
                    if current_sheet_rows >= MAX_EXCEL_ROWS:
                        sheet_idx += 1
                        worksheet = workbook.add_worksheet(f"Sheet{sheet_idx}")
                        worksheet.write_row(0, 0, headers)
                        current_sheet_rows = 0
                        self.msg_queue.put({'type': 'status', 'message': f"当前工作表已满，已创建 Sheet{sheet_idx}..."})
                        
                    worksheet.write_row(current_sheet_rows + 1, 0, row)
                    current_sheet_rows += 1
                    total_rows_written += 1
                    
                self.msg_queue.put({'type': 'status', 'message': f"已处理并写入数据：共 {total_rows_written:,} 行"})
                
            if worksheet is None:
                worksheet = workbook.add_worksheet("Sheet1")
                
            self.msg_queue.put({'type': 'status', 'message': f"正在压缩并保存 XLSX 文件..."})
            workbook.close()
            
            elapsed = time.time() - start_time
            msg = f"转换完成！已保存为：{os.path.basename(excel_file)}\n共写入 {total_rows_written:,} 行，耗时: {elapsed:.2f}秒"
            self.msg_queue.put({'type': 'status', 'message': msg})
            
            self.root.after(0, lambda: messagebox.showinfo("成功", f"CSV 转换 Excel 成功！\n保存至：{excel_file}\n耗时: {elapsed:.2f}秒"))
            self.root.after(0, lambda: self.file_path.set(excel_file))
            self.root.after(0, self.update_file_type)
            
        except Exception as e:
            error_msg = f"CSV 转换失败: {str(e)}"
            self.logger.error(error_msg)
            self.msg_queue.put({'type': 'error', 'message': error_msg})
            self.root.after(0, lambda: messagebox.showerror("错误", f"CSV 转换失败：{str(e)}"))
        finally:
            self.root.after(0, lambda: self.set_buttons_state(True))

    def save_processed_raw_xlsx(self):
        """将读取并清洗后的 FLOEFD 原始数据保存为 xlsx 文件"""
        if self.result_df is None:
            messagebox.showwarning("警告", "当前无有效数据，请先载入并读取数据！")
            return
        
        file_path = self.file_path.get()
        if not file_path:
            messagebox.showwarning("警告", "未获取到输入文件路径！")
            return
            
        try:
            save_path = os.path.splitext(file_path)[0] + '_processed.xlsx'
            self.set_buttons_state(False)
            self.update_status(f"正在后台保存清洗后的 Excel 文件到: {save_path}...")
            
            threading.Thread(
                target=self._bg_save_processed_xlsx,
                args=(save_path,),
                daemon=True
            ).start()
        except Exception as e:
            messagebox.showerror("错误", f"启动保存线程失败: {str(e)}")
            self.set_buttons_state(True)

    def _bg_save_processed_xlsx(self, save_path):
        try:
            # 使用 xlsxwriter 引擎保存 DataFrame 到 Excel 格式
            with pd.ExcelWriter(save_path, engine='xlsxwriter') as writer:
                self.result_df.to_excel(writer, sheet_name='Sheet1', index=False)
            
            self.msg_queue.put({
                'type': 'status',
                'message': f"数据处理后成功保存到：{save_path}"
            })
            self.root.after(0, lambda: messagebox.showinfo("成功", f"数据已成功保存至：\n{os.path.basename(save_path)}"))
        except Exception as e:
            error_msg = f"保存 Excel 失败: {str(e)}"
            self.logger.error(error_msg)
            self.msg_queue.put({'type': 'error', 'message': error_msg})
            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
        finally:
            self.root.after(0, lambda: self.set_buttons_state(True))
