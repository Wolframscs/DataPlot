import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import os
import time
import gc
import logging
import json
import openpyxl
import sys
import threading
import queue

try:
    import python_calamine
    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False
# 设置matplotlib支持中文显示
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 用来正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

# 全局配置 Matplotlib 性能优化参数
import matplotlib
matplotlib.rcParams['path.simplify'] = True
matplotlib.rcParams['path.simplify_threshold'] = 1.0
matplotlib.rcParams['agg.path.chunksize'] = 10000

def resource_path(relative_path):
    """ 获取资源的绝对路径，兼容开发环境与PyInstaller打包后的环境 """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class CustomNavigationToolbar(NavigationToolbar2Tk):
    def __init__(self, canvas, parent):
        # 创建一个框架来容纳工具栏
        self.toolbar_frame = ttk.Frame(parent)
        self.toolbar_frame.grid(row=1, column=0, sticky='ew')
        super().__init__(canvas, self.toolbar_frame)

    def _Button(self, text, image_file, toggle, command):
        # 重写按钮创建方法，使用grid而不是pack
        b = super()._Button(text, image_file, toggle, command)
        b.pack_forget()  # 取消pack布局
        b.grid(row=0, column=len(self.toolbar_frame.children)-1, padx=2, pady=2)  # 使用grid布局
        return b

    def _Spacer(self):
        # 重写分隔符创建方法，使用grid而不是pack
        s = super()._Spacer()
        s.pack_forget()  # 取消pack布局
        s.grid(row=0, column=len(self.toolbar_frame.children)-1, padx=4)  # 使用grid布局
        return s

class PlotterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("DataPlot")
        try:
            icon_path = resource_path('icon.ico')
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception:
            pass
        self.result_df = None
        self.CHUNK_SIZE = 100000  # 定义为类属性
        self.msg_queue = queue.Queue()
        
        # 设置日志
        logging.basicConfig(
            filename='DataPlot.log',
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger('DataPlot')
        
        # 添加窗口关闭事件处理
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        try:
            # 定义线型选项
            self.line_styles_dict = {
                '实线': '-',
                '虚线': '--',
                '点线': ':',
                '点划线': '-.',
            }
            
            # 初始化线型列表
            self.line_styles = []
            
            # 获取屏幕分辨率，并设置初始窗口几何大小与位置 (居中显示)
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            
            # 计算缩放比例因子（以 1920x1080 为基准）
            scale_factor = min(screen_width / 1920.0, screen_height / 1080.0)
            scale_factor = max(0.8, min(scale_factor, 2.0))  # 限制缩放因子范围在 0.8 到 2.0 之间
            
            # 根据分辨率计算初始窗口大小 (宽高各占 75%)
            window_width = int(screen_width * 0.9)
            window_height = int(screen_height * 0.75)
            
            # 限制窗口范围，避免溢出或过小
            window_width = max(1100, min(window_width, screen_width - 100))
            window_height = max(750, min(window_height, screen_height - 100))
            
            # 居中坐标
            x_offset = (screen_width - window_width) // 2
            y_offset = (screen_height - window_height) // 2
            self.root.geometry(f"{window_width}x{window_height}+{x_offset}+{y_offset}")
            

            
            # 动态计算 GUI 默认字体大小 (基准值为 12)
            self.GUI_FONT_SIZE = max(10, min(18, int(12 * scale_factor)))
            default_font = ('Microsoft YaHei', self.GUI_FONT_SIZE)
            
            # 动态计算 Matplotlib 图表大小
            self.fig_w = max(8.0, min(16.0, 12.0 * scale_factor))
            self.fig_h = max(5.0, min(11.0, 8.0 * scale_factor))

            # 配置根窗口 of Grid Weight，使其可以跟随窗口调整大小
            self.root.grid_rowconfigure(0, weight=1)
            self.root.grid_columnconfigure(0, weight=1)

            
            # 定义统一的组件宽度和位置常量
            self.LABEL_WIDTH = 8       # 左侧标签宽度
            self.SMALL_LABEL_WIDTH = 6 # 小标签宽度（"最小值"/"最大值"等标签）
            self.Y_LABEL_WIDTH = 6     # Y1/Y2/Y3标签宽度
            self.ENTRY_WIDTH = 8       # 输入框宽度
            self.COMBO_WIDTH = 6       # 下拉框宽度（减小以补偿下拉箭头的宽度）
            self.TITLE_ENTRY_WIDTH = 15 # 标题输入框宽度
            self.SETTING_LABEL_WIDTH = 8  # 设置标签的统一宽度
            self.LABEL_PADX = (8, 0)  # 统一的标签左边距
            self.WIDGET_PADX = 1       # 组件之间的间距
            
            # 设置按钮样式
            style = ttk.Style()
            style.configure('Big.TButton', font=default_font)
            style.configure('Custom.TLabelframe.Label', font=default_font)
            style.configure('Custom.TCheckbutton', font=default_font)  # 添加 Checkbutton 样式
            
            # 应用默认字体
            self.root.option_add('*Font', default_font)
            
            # 创建统一的标签样式
            label_style = {'width': 10, 'anchor': tk.W}  # 减小宽度，使标签更紧凑
            
            # 添加在类初始化时
            self.version = "1.0.0"
            self._update_timer = None  # 用于延迟更新
            self._last_plot_time = 0   # 用于限制更新频率
            self._is_loading_settings = False
            self.auto_downsample = tk.BooleanVar(value=True)
            self.max_plot_points = tk.StringVar(value="10000")
            
            # 创建主框架
            self.main_frame = ttk.Frame(root, padding="10")
            self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            
            # 配置主框架的网格权重
            self.main_frame.grid_rowconfigure(0, weight=1)
            self.main_frame.grid_columnconfigure(1, weight=1)  # 图表区域自动填充并缩放剩余空间
            self.main_frame.grid_columnconfigure(0, weight=0)  # 控制面板列不拉伸，保留所需尺寸防止被遮挡或压缩
            
            # 创建左侧控制面板框架
            control_frame = ttk.Frame(self.main_frame)
            control_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5)
            
            # 设置LabelFrame的样式
            style = ttk.Style()
            style.configure('Custom.TLabelframe.Label', font=('Microsoft YaHei', 12))
            
            # 在文件选择部分添加文件类型选择
            file_type_frame = ttk.Frame(control_frame)
            file_type_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
            
            # 文件类型选择
            ttk.Label(file_type_frame, text="文件类型:", **label_style).grid(row=0, column=0, sticky=tk.W)
            self.file_type = tk.StringVar(value="raw")
            
            # 使用 tk.Radiobutton 代替 ttk.Radiobutton
            tk.Radiobutton(file_type_frame, text="原始Excel", variable=self.file_type, value="raw", 
                          command=self.update_file_type,
                          font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=0, column=1)
            tk.Radiobutton(file_type_frame, text="后处理文件", variable=self.file_type, value="processed", 
                          command=self.update_file_type,
                          font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=0, column=2)
            tk.Radiobutton(file_type_frame, text="电池数据", variable=self.file_type, value="battery", 
                          command=self.update_file_type,
                          font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=0, column=3)
            
            # 文件路径选择
            file_frame = ttk.Frame(control_frame)
            file_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
            ttk.Label(file_frame, text="文件路径:", **label_style).grid(row=0, column=0, sticky=tk.W)
            self.file_path = tk.StringVar()
            ttk.Entry(file_frame, textvariable=self.file_path, width=30,  # 统一宽度为30
                     font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=0, column=1, padx=10, sticky=tk.W)
            self.browse_btn = ttk.Button(file_frame, text="浏览", command=self.browse_file,
                      style='Big.TButton', width=10)
            self.browse_btn.grid(row=0, column=2)
            
            # Sheet选择
            sheet_frame = ttk.Frame(control_frame)
            sheet_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
            ttk.Label(sheet_frame, text="表格名称:", **label_style).grid(row=0, column=0, sticky=tk.W)
            self.sheet_name = tk.StringVar()
            self.sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_name,
                                          width=30,
                                          state='readonly', 
                                          font=('Microsoft YaHei', self.GUI_FONT_SIZE))
            self.sheet_combo.grid(row=0, column=1, sticky=tk.W, padx=10)
            
            # CSV转Excel按钮，放置在表格名称 Combobox 右侧，正好在“浏览”按钮正下方
            self.csv2xlsx_btn = ttk.Button(sheet_frame, text="csv2xlsx", command=self.convert_csv_to_xlsx,
                                           style='Big.TButton', width=10)
            self.csv2xlsx_btn.grid(row=0, column=2)
            
            # 添加起始行输入框
            start_row_frame = ttk.Frame(control_frame)
            start_row_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
            ttk.Label(start_row_frame, text="起始行:", **label_style).grid(row=0, column=0, sticky=tk.W)
            self.start_row = tk.StringVar(value="1")  # 默认从第1行开始
            ttk.Entry(start_row_frame, textvariable=self.start_row, width=10).grid(row=0, column=1, sticky=tk.W, padx=10)
            
            # 计算按钮移到新的一行
            self.process_btn = ttk.Button(start_row_frame, text="读取", command=self.process_data,
                      style='Big.TButton', width=10)
            self.process_btn.grid(row=0, column=2, padx=(0, 5))
            
            # 创建电池数据配置框架 (默认不显示)
            self.battery_filter_frame = ttk.LabelFrame(control_frame, text="电池数据配置", padding="5", style='Custom.TLabelframe')
            
            self.cycle_col = tk.StringVar()
            self.step_col = tk.StringVar()
            self.time_col = tk.StringVar()
            
            # 第一行：列名映射
            ttk.Label(self.battery_filter_frame, text="循环列:", font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=0, column=0, sticky=tk.W, padx=5)
            self.cycle_col_combo = ttk.Combobox(self.battery_filter_frame, textvariable=self.cycle_col, state='readonly', font=('Microsoft YaHei', self.GUI_FONT_SIZE), width=10)
            self.cycle_col_combo.grid(row=0, column=1, sticky=tk.W, padx=5)
            self.cycle_col_combo.bind('<<ComboboxSelected>>', self.on_cycle_col_changed)
            
            ttk.Label(self.battery_filter_frame, text="工步列:", font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=0, column=2, sticky=tk.W, padx=5)
            self.step_col_combo = ttk.Combobox(self.battery_filter_frame, textvariable=self.step_col, state='readonly', font=('Microsoft YaHei', self.GUI_FONT_SIZE), width=10)
            self.step_col_combo.grid(row=0, column=3, sticky=tk.W, padx=5)
            self.step_col_combo.bind('<<ComboboxSelected>>', self.on_step_col_changed)
            
            ttk.Label(self.battery_filter_frame, text="时间列:", font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=0, column=4, sticky=tk.W, padx=5)
            self.time_col_combo = ttk.Combobox(self.battery_filter_frame, textvariable=self.time_col, state='readonly', font=('Microsoft YaHei', self.GUI_FONT_SIZE), width=10)
            self.time_col_combo.grid(row=0, column=5, sticky=tk.W, padx=5)
            self.time_col_combo.bind('<<ComboboxSelected>>', self.on_time_col_changed)

            # 第二行：筛选条件
            ttk.Label(self.battery_filter_frame, text="循环筛选:", font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=1, column=0, sticky=tk.W, padx=5)
            self.cycle_filter = tk.StringVar(value="全部")
            self.cycle_filter_combo = ttk.Combobox(self.battery_filter_frame, textvariable=self.cycle_filter, state='readonly', font=('Microsoft YaHei', self.GUI_FONT_SIZE), width=10)
            self.cycle_filter_combo.grid(row=1, column=1, sticky=tk.W, padx=5)
            
            ttk.Label(self.battery_filter_frame, text="工步筛选:", font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=1, column=2, sticky=tk.W, padx=5)
            self.step_filter = tk.StringVar(value="全部")
            self.step_filter_combo = ttk.Combobox(self.battery_filter_frame, textvariable=self.step_filter, state='readonly', font=('Microsoft YaHei', self.GUI_FONT_SIZE), width=10)
            self.step_filter_combo.grid(row=1, column=3, sticky=tk.W, padx=5)
            
            # 计算按钮 (点击后更新图表)
            self.battery_calc_btn = ttk.Button(self.battery_filter_frame, text="计算", command=self.delayed_update,
                       style='Big.TButton', width=10)
            self.battery_calc_btn.grid(row=1, column=4, columnspan=2, sticky=tk.W, padx=15)
            
            # 创建单Y轴绘图选项框架
            plot_frame = ttk.LabelFrame(control_frame, text="绘图选项", padding="5", style='Custom.TLabelframe')
            plot_frame.grid(row=5, column=0, columnspan=3, pady=5, sticky=(tk.W, tk.E))
            
            # X轴选择
            ttk.Label(plot_frame, text="X轴:", 
                     font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=0, column=0, sticky=tk.W)
            self.x_axis = tk.StringVar()
            self.x_combo = ttk.Combobox(plot_frame, textvariable=self.x_axis, 
                                      state='readonly', font=('Microsoft YaHei', self.GUI_FONT_SIZE),
                                      width=15)  # 设置固定宽度为15
            self.x_combo.grid(row=0, column=1, sticky=tk.W, padx=5)  # 改为左对齐
            
            # X轴标题
            ttk.Label(plot_frame, text="标题:").grid(row=0, column=2, padx=(10,0), sticky=tk.W)
            self.x_title = ttk.Entry(plot_frame, width=20)
            self.x_title.grid(row=0, column=3, padx=2, sticky=tk.W)
            self.x_title.insert(0, "Time/s")  # 默认值
            
            # Y轴选择和按钮
            y_frame = ttk.Frame(plot_frame)
            y_frame.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E), padx=5)  # 扩展到4列
            ttk.Label(y_frame, text="Y轴:", 
                     font=('Microsoft YaHei', self.GUI_FONT_SIZE)).grid(row=0, column=0, sticky=tk.W)
            
            # 创建三个Listbox
            self.y_listboxes = []
            self.y_selections = [[], [], []]  # 保存每个Y轴的选择状态
            
            for i in range(3):
                frame = ttk.Frame(y_frame)
                frame.grid(row=0, column=i+1, padx=2)
                
                ttk.Label(frame, text=f"Y{i+1}:", 
                         font=('Microsoft YaHei', self.GUI_FONT_SIZE)
                         ).grid(row=0, column=0, sticky=tk.W)
                
                listbox = tk.Listbox(frame, selectmode=tk.EXTENDED, height=5, width=15,  # 统一宽度为15
                                   exportselection=False,
                                   font=('Microsoft YaHei', self.GUI_FONT_SIZE))
                listbox.grid(row=1, column=0)
                listbox.bind('<<ListboxSelect>>', self.on_selection_change)
                self.y_listboxes.append(listbox)
            
            # 按钮框架
            button_frame = ttk.Frame(y_frame)
            button_frame.grid(row=0, column=4, padx=5)
            
            self.clear_btn = ttk.Button(button_frame, text="清除选择", 
                      command=self.clear_all_selections,
                      style='Big.TButton', width=10)
            self.clear_btn.grid(row=0, column=0, pady=2)
            
            self.select_all_btn = ttk.Button(button_frame, text="全部选择", 
                      command=self.select_all_y,
                      style='Big.TButton', width=10)
            self.select_all_btn.grid(row=1, column=0, pady=2)
            
            self.plot_btn = ttk.Button(button_frame, text="绘制图表", 
                      command=self.delayed_update,
                      style='Big.TButton', width=10)
            self.plot_btn.grid(row=2, column=0, pady=2)
            
            self.save_btn = ttk.Button(button_frame, text="保存数据", 
                      command=self.save_plot_data,
                      style='Big.TButton', width=10)
            self.save_btn.grid(row=3, column=0, pady=2)
            
            # 图例设置
            legend_frame = ttk.Frame(plot_frame)
            legend_frame.grid(row=3, column=0, columnspan=4, sticky=(tk.W, tk.E), padx=5)
            ttk.Label(legend_frame, text="图例设置:", width=self.SETTING_LABEL_WIDTH).grid(
                row=0, column=0, sticky=(tk.W, tk.E))
            # 垂直位置控制
            ttk.Label(legend_frame, text="垂直:", width=self.Y_LABEL_WIDTH).grid(
                row=0, column=1, sticky=(tk.W, tk.E), padx=self.LABEL_PADX)
            self.legend_y = tk.StringVar(value="1.02")
            ttk.Entry(legend_frame, textvariable=self.legend_y, width=self.ENTRY_WIDTH).grid(
                row=0, column=2, sticky=(tk.W, tk.E), padx=self.WIDGET_PADX)

            # 水平位置控制（单个输入框）
            ttk.Label(legend_frame, text="水平:", width=self.Y_LABEL_WIDTH).grid(
                row=0, column=3, sticky=(tk.W, tk.E), padx=self.LABEL_PADX)
            self.legend_x_positions_str = tk.StringVar(value="0, 0.3, 0.7")
            ttk.Entry(legend_frame, textvariable=self.legend_x_positions_str, width=self.ENTRY_WIDTH).grid(
                row=0, column=4, sticky=(tk.W, tk.E), padx=self.WIDGET_PADX)
            self.legend_x_positions_str.trace_add('write', lambda *args: self.update_legend_positions())

            # 添加图例显示/隐藏控制
            self.legend_visible = tk.BooleanVar(value=True)
            ttk.Checkbutton(legend_frame, text="显示图例", 
                          variable=self.legend_visible,
                          command=self.toggle_legend,
                          style='Custom.TCheckbutton').grid(
                row=0, column=5, padx=(10, 0), sticky=tk.W)

            # Y轴范围设置和标题
            y_ranges_frame = ttk.Frame(plot_frame)
            y_ranges_frame.grid(row=4, column=0, columnspan=4, sticky=(tk.W, tk.E), padx=5)
            
            # Y轴设置
            self.y_settings = []
            default_y_configs = [
                {'min': '20', 'max': '60', 'title': 'Temperature/℃'},
                {'min': '0', 'max': '500', 'title': 'Current/A'},
                {'min': '2', 'max': '4', 'title': 'Voltage/V'}
            ]
            for i in range(3):
                config = default_y_configs[i]
                settings = {
                    'min': tk.StringVar(value=config['min']),
                    'max': tk.StringVar(value=config['max']),
                    'title': tk.StringVar(value=config['title'])
                }
                self.y_settings.append(settings)
                
                # Y轴范围和标题设置
                ttk.Label(y_ranges_frame, text=f"Y{i+1}轴范围:", width=self.LABEL_WIDTH).grid(
                    row=i, column=0, sticky=tk.W)
                ttk.Label(y_ranges_frame, text="最小值:", width=self.SMALL_LABEL_WIDTH).grid(
                    row=i, column=1, sticky=tk.W, padx=self.LABEL_PADX)
                ttk.Entry(y_ranges_frame, textvariable=self.y_settings[i]['min'], 
                         width=self.ENTRY_WIDTH).grid(row=i, column=2, sticky=tk.W, padx=self.WIDGET_PADX)
                ttk.Label(y_ranges_frame, text="最大值:", width=self.SMALL_LABEL_WIDTH).grid(
                    row=i, column=3, sticky=tk.W, padx=self.LABEL_PADX)
                ttk.Entry(y_ranges_frame, textvariable=self.y_settings[i]['max'], 
                         width=self.ENTRY_WIDTH).grid(row=i, column=4, sticky=tk.W, padx=self.WIDGET_PADX)
                ttk.Label(y_ranges_frame, text="标题:", width=self.SMALL_LABEL_WIDTH).grid(
                    row=i, column=5, sticky=tk.W, padx=self.LABEL_PADX)
                ttk.Entry(y_ranges_frame, textvariable=self.y_settings[i]['title'], 
                         width=self.TITLE_ENTRY_WIDTH).grid(row=i, column=6, sticky=tk.W, padx=self.WIDGET_PADX)
                
                # 绑定更新事件，使用lambda捕获当前的i值
                self.y_settings[i]['min'].trace_add('write', lambda *args, axis=i: self.update_y_axis(axis))
                self.y_settings[i]['max'].trace_add('write', lambda *args, axis=i: self.update_y_axis(axis))
                self.y_settings[i]['title'].trace_add('write', lambda *args, axis=i: self.update_y_axis(axis))

            # 添加空行
            ttk.Frame(plot_frame, height=10).grid(row=5, column=0, columnspan=4, pady=5)

            # 字体设置
            font_frame = ttk.Frame(plot_frame)
            font_frame.grid(row=6, column=0, columnspan=4, sticky=(tk.W, tk.E), padx=5)
            ttk.Label(font_frame, text="绘图字体:", width=self.SETTING_LABEL_WIDTH).grid(
                row=0, column=0, sticky=(tk.W, tk.E))
            
            ttk.Label(font_frame, text="字体:", width=self.Y_LABEL_WIDTH).grid(
                row=0, column=1, sticky=(tk.W, tk.E), padx=self.LABEL_PADX)
            self.font_family = tk.StringVar(value="SimHei")  # 使用黑体作为默认字体
            font_combo = ttk.Combobox(font_frame, textvariable=self.font_family,
                                    values=["SimHei", "SimSun", "KaiTi", "FangSong", "Arial"],
                                    state='readonly', width=self.COMBO_WIDTH)
            font_combo.grid(row=0, column=2, padx=self.WIDGET_PADX)

            ttk.Label(font_frame, text="大小:", width=self.Y_LABEL_WIDTH).grid(
                row=0, column=3, sticky=(tk.W, tk.E), padx=self.LABEL_PADX)
            self.font_size = tk.StringVar(value="15")  # 默认值改为15
            # 创建8到24的字体大小列表
            font_sizes = [str(size) for size in range(8, 25, 1)]
            font_size_combo = ttk.Combobox(font_frame, textvariable=self.font_size,
                                         values=font_sizes,
                                         state='readonly', width=self.COMBO_WIDTH)
            font_size_combo.grid(row=0, column=4, sticky=(tk.W, tk.E), padx=self.WIDGET_PADX)

            # 线型设置（包含线型宽度和线型）
            line_style_frame = ttk.Frame(plot_frame)
            line_style_frame.grid(row=7, column=0, columnspan=4, sticky=(tk.W, tk.E), padx=5)
            ttk.Label(line_style_frame, text="线型宽度:", width=self.SETTING_LABEL_WIDTH).grid(
                row=0, column=0, sticky=tk.W)

            # 框线和曲线宽度设置
            ttk.Label(line_style_frame, text="框线:", width=self.Y_LABEL_WIDTH).grid(
                row=0, column=2*0+1, sticky=tk.W, padx=self.LABEL_PADX)
            self.frame_width = tk.StringVar(value="1.5")
            frame_width_combo = ttk.Combobox(line_style_frame, textvariable=self.frame_width,
                                           values=["0.5", "1.0", "1.5", "2.0"],
                                           state='readonly', width=self.COMBO_WIDTH)
            frame_width_combo.grid(row=0, column=2*0+2, padx=self.WIDGET_PADX)

            ttk.Label(line_style_frame, text="曲线:", width=self.Y_LABEL_WIDTH).grid(
                row=0, column=2*1+1, sticky=tk.W, padx=self.LABEL_PADX)
            self.line_width = tk.StringVar(value="1.5")
            line_width_combo = ttk.Combobox(line_style_frame, textvariable=self.line_width,
                                          values=["0.5", "1.0", "1.5", "2.0"],
                                          state='readonly', width=self.COMBO_WIDTH)
            line_width_combo.grid(row=0, column=2*1+2, padx=self.WIDGET_PADX)

            # 线型设置
            ttk.Label(line_style_frame, text="线型类型:", width=self.SETTING_LABEL_WIDTH).grid(
                row=1, column=0, sticky=tk.W)

            # Y1/Y2/Y3线型选择
            for i in range(3):
                ttk.Label(line_style_frame, text=f"Y{i+1}:", width=self.Y_LABEL_WIDTH).grid(
                    row=1, column=2*i+1, sticky=tk.W, padx=self.LABEL_PADX)
                default_style = '点划线' if i == 2 else list(self.line_styles_dict.keys())[i]
                style_var = tk.StringVar(value=default_style)
                style_combo = ttk.Combobox(line_style_frame, textvariable=style_var,
                                         values=list(self.line_styles_dict.keys()),
                                         state='readonly', width=self.COMBO_WIDTH)
                style_combo.grid(row=1, column=2*i+2, padx=self.WIDGET_PADX)
                self.line_styles.append(style_var)
                style_var.trace_add('write', lambda *args: self.update_plot())

            # 配色方案设置
            color_scheme_frame = ttk.Frame(plot_frame)
            color_scheme_frame.grid(row=8, column=0, columnspan=4, sticky=(tk.W, tk.E), padx=5)
            ttk.Label(color_scheme_frame, text="绘图配色:", width=self.SETTING_LABEL_WIDTH).grid(
                row=0, column=0, sticky=tk.W)

            # 定义颜色方案选项
            self.color_schemes_dict = {
                '默认': None,
                'Set1': plt.cm.Set1,
                'Set2': plt.cm.Set2,
                'Set3': plt.cm.Set3,
                'Paired': plt.cm.Paired,
                'Dark2': plt.cm.Dark2,
                'Accent': plt.cm.Accent,
                'Pastel1': plt.cm.Pastel1,
                'Pastel2': plt.cm.Pastel2
            }

            # 添加配色方案设置
            self.color_schemes = []
            for i in range(3):
                ttk.Label(color_scheme_frame, text=f"Y{i+1}:", width=self.Y_LABEL_WIDTH).grid(
                    row=0, column=2*i+1, sticky=tk.W, padx=self.LABEL_PADX)
                # 为Y3设置默认Dark2，其他保持不变
                default_scheme = 'Dark2' if i == 2 else list(self.color_schemes_dict.keys())[i]
                scheme_var = tk.StringVar(value=default_scheme)
                scheme_combo = ttk.Combobox(color_scheme_frame, textvariable=scheme_var,
                                          values=list(self.color_schemes_dict.keys()),
                                          state='readonly', width=self.COMBO_WIDTH)
                scheme_combo.grid(row=0, column=2*i+2, padx=self.WIDGET_PADX)
                self.color_schemes.append(scheme_var)
                scheme_var.trace_add('write', lambda *args: self.update_plot())

            # 性能优化设置
            perf_frame = ttk.Frame(plot_frame)
            perf_frame.grid(row=9, column=0, columnspan=4, sticky=(tk.W, tk.E), padx=5)
            ttk.Label(perf_frame, text="性能优化:", width=self.SETTING_LABEL_WIDTH).grid(
                row=0, column=0, sticky=tk.W)
            
            self.auto_downsample_cb = ttk.Checkbutton(perf_frame, text="自动降采样", 
                                                      variable=self.auto_downsample,
                                                      command=self.update_plot,
                                                      style='Custom.TCheckbutton')
            self.auto_downsample_cb.grid(row=0, column=1, padx=self.LABEL_PADX, sticky=tk.W)
            
            ttk.Label(perf_frame, text="最大点数:", width=self.Y_LABEL_WIDTH).grid(
                row=0, column=2, sticky=tk.W, padx=self.LABEL_PADX)
            self.max_plot_points_combo = ttk.Combobox(perf_frame, textvariable=self.max_plot_points,
                                                      values=["50000", "100000", "200000", "500000", "无限制"],
                                                      state='readonly', width=self.COMBO_WIDTH)
            self.max_plot_points_combo.grid(row=0, column=3, padx=self.WIDGET_PADX, sticky=tk.W)
            self.max_plot_points_combo.bind('<<ComboboxSelected>>', lambda e: self.update_plot())

            # 将状态信息显示区域移动到配色方案下方
            status_frame = ttk.Frame(control_frame)  # 改为使用control_frame作为父框架
            status_frame.grid(row=10, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)  # 放在配色方案和性能优化下方
            
            # 创建状态信息文本框
            self.status_text = tk.Text(status_frame, height=6, width=40,  # 调整高度和宽度
                                     font=default_font)
            self.status_text.grid(row=0, column=0, sticky=(tk.W, tk.E))
            
            # 添加滚动条
            scrollbar = ttk.Scrollbar(status_frame, orient="vertical", 
                                    command=self.status_text.yview)
            scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
            self.status_text.configure(yscrollcommand=scrollbar.set)
            
            # 配置status_frame的列权重
            status_frame.grid_columnconfigure(0, weight=1)

            # 创建图表框架
            plot_display_frame = ttk.Frame(self.main_frame)
            plot_display_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5)
            
            # 配置图表框架的网格权重
            plot_display_frame.grid_rowconfigure(0, weight=1)
            plot_display_frame.grid_columnconfigure(0, weight=1)
            
            # 创建图表区域
            canvas_frame = ttk.Frame(plot_display_frame)
            canvas_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            canvas_frame.grid_rowconfigure(0, weight=1)
            canvas_frame.grid_columnconfigure(0, weight=1)

            # 图表显示区域
            self.fig = plt.figure(figsize=(self.fig_w, self.fig_h))
            self.ax = self.fig.add_subplot(111)
            self.canvas = FigureCanvasTkAgg(self.fig, master=canvas_frame)
            self.canvas.draw()
            self.canvas.get_tk_widget().grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

            # 创建工具栏
            toolbar_frame = ttk.Frame(canvas_frame)
            toolbar_frame.grid(row=1, column=0, sticky=(tk.W, tk.E))
            self.toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
            self.toolbar.update()
            
            # 初始化Y轴对象
            self.axes = {'y1': self.ax}  # 使用字典存储轴对象
            self.current_axes = {'y1': self.ax}  # 当前活动的轴
            
            # 绑定窗口大小改变事件，使用 draw_idle 响应调整，防止主线程死锁
            self.root.bind('<Configure>', self.on_window_resize)
            
            # 只保留这些绑定
            self.x_title.bind('<Return>', lambda e: self.update_plot())
            
            # 根据分辨率限制窗口最小尺寸
            min_width = min(1200, int(window_width * 0.9))
            min_height = min(800, int(window_height * 0.85))
            self.root.minsize(min_width, min_height)


            # 绑定事件
            self.font_family.trace_add('write', lambda *args: self.update_font_and_plot())
            self.font_size.trace_add('write', lambda *args: self.update_font_and_plot())
            self.legend_y.trace_add('write', lambda *args: self.update_legend_only())

            # 自动加载设置
            self.load_settings()
            
            # 启动轮询队列
            self.check_queue()

            self.logger.info("应用程序启动成功")
        except Exception as e:
            self.logger.error(f"初始化失败: {str(e)}")
            raise

    def check_queue(self):
        """检查消息队列以在主线程中安全地更新 UI"""
        try:
            while True:
                msg = self.msg_queue.get_nowait()
                msg_type = msg.get('type')
                if msg_type == 'status':
                    message = msg.get('message')
                    clear = msg.get('clear', False)
                    # 直接在主线程文本框插入日志以防递归
                    if clear:
                        self.status_text.delete(1.0, tk.END)
                    self.status_text.insert(tk.END, message + "\n")
                    self.status_text.see(tk.END)
                elif msg_type == 'done':
                    df = msg.get('df')
                    self.result_df = df
                    
                    # 更新下拉菜单可选列
                    cols_list = list(df.columns)
                    self.cycle_col_combo['values'] = cols_list
                    self.step_col_combo['values'] = cols_list
                    self.time_col_combo['values'] = cols_list
                    
                    self.cycle_col.set(msg.get('cycle_default', ''))
                    self.step_col.set(msg.get('step_default', ''))
                    self.time_col.set(msg.get('time_default', ''))
                    
                    self.cycle_filter_combo['values'] = msg.get('cycles', ['全部'])
                    self.cycle_filter_combo.set('全部')
                    
                    self.step_filter_combo['values'] = msg.get('steps', ['全部'])
                    self.step_filter_combo.set('全部')
                    
                    self.update_listboxes()
                    self.update_status(msg.get('message', "文件读取和处理完成"))
                    self.set_buttons_state(True)
                elif msg_type == 'raw_done':
                    df = msg.get('df')
                    self.result_df = df
                    self.update_listboxes()
                    self.update_status(msg.get('message', "数据处理完成"))
                    self.set_buttons_state(True)
                elif msg_type == 'time_diff_done':
                    self.update_listboxes()
                    self.update_status(msg.get('message'))
                    self.set_buttons_state(True)
                    self.delayed_update()
                elif msg_type == 'error':
                    error_msg = msg.get('message')
                    self.update_status(error_msg, clear=True)
                    self.set_buttons_state(True)
                self.msg_queue.task_done()
        except queue.Empty:
            pass
        self.root.after(100, self.check_queue)

    def set_buttons_state(self, enabled=True):
        """控制交互按钮状态与光标形状，避免重复并发操作"""
        state = 'normal' if enabled else 'disabled'
        self.process_btn.configure(state=state)
        self.browse_btn.configure(state=state)
        self.battery_calc_btn.configure(state=state)
        self.clear_btn.configure(state=state)
        self.select_all_btn.configure(state=state)
        self.plot_btn.configure(state=state)
        self.save_btn.configure(state=state)
        self.csv2xlsx_btn.configure(state=state)
        
        # 改变指针光标形状以指示繁忙状态
        cursor = "" if enabled else "watch"
        self.root.configure(cursor=cursor)

    def delayed_update(self, *args):
        """延迟更新以减少重绘频率"""
        if hasattr(self, '_update_timer'):
            try:
                self.root.after_cancel(self._update_timer)
            except Exception:
                pass
        self._update_timer = self.root.after(200, self.update_plot)

    def update_file_type(self):
        # 更新文件选择对话框的文件类型
        if self.file_type.get() == "battery":
            self.battery_filter_frame.grid(row=4, column=0, columnspan=3, pady=5, sticky=(tk.W, tk.E))
        else:
            self.battery_filter_frame.grid_forget()

        if self.file_type.get() == "raw":
            self.sheet_combo.configure(state='readonly')
        else:
            if self.file_path.get().endswith('.xlsx'):
                self.sheet_combo.configure(state='readonly')
            else:
                self.sheet_combo.configure(state='disabled')

    def browse_file(self):
        if self.file_type.get() == "raw":
            filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]
        else:
            filetypes = [("Excel/CSV files", "*.xlsx;*.csv"), ("Excel files", "*.xlsx"), 
                        ("CSV files", "*.csv"), ("All files", "*.*")]
            
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename:
            self.file_path.set(filename)
            try:
                if filename.endswith('.xlsx'):
                    # 更新sheet列表
                    engine = 'calamine' if HAS_CALAMINE else None
                    excel_file = pd.ExcelFile(filename, engine=engine)
                    self.sheet_combo['values'] = excel_file.sheet_names
                    if excel_file.sheet_names:
                        self.sheet_combo.set(excel_file.sheet_names[0])
                    self.sheet_combo.configure(state='readonly')
                else:
                    self.sheet_combo.configure(state='disabled')
            except Exception as e:
                messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")

    def update_y_options(self):
        """更新Y轴选项"""
        # 获取所有列名
        columns = self.result_df.columns.tolist()
        
        # 更新X轴下拉框
        self.x_combo['values'] = columns
        self.x_combo.set(columns[0])  # 默认选择第一列
        
        # 清除并更新所有Y轴列表框
        for listbox in self.y_listboxes:
            listbox.delete(0, tk.END)
            # 添加除X轴外的所有列作为Y轴选项
            for col in columns[1:]:
                listbox.insert(tk.END, col)

    def process_data(self):
        """处理Excel/CSV数据"""
        try:
            file_path = self.file_path.get()
            if not os.path.exists(file_path):
                self.update_status("错误：文件不存在", clear=True)
                return
                
            self.set_buttons_state(False)
            
            file_type = self.file_type.get()
            sheet_name = self.sheet_name.get()
            start_row_str = self.start_row.get()
            
            threading.Thread(
                target=self._bg_process_data,
                args=(file_path, file_type, sheet_name, start_row_str),
                daemon=True
            ).start()
            
        except Exception as e:
            error_msg = f"启动处理线程失败: {str(e)}"
            self.logger.error(error_msg)
            self.update_status(error_msg, clear=True)
            self.set_buttons_state(True)

    def _bg_process_data(self, file_path, file_type, sheet_name, start_row_str):
        """后台线程处理函数，避免阻塞 UI 主线程"""
        try:
            if file_type == "raw":
                self.msg_queue.put({'type': 'status', 'message': "开始读取原始Excel文件...", 'clear': True})
                try:
                    # 记录开始时间
                    start_time = time.time()
                    
                    # 直接读取Excel文件
                    self.msg_queue.put({'type': 'status', 'message': "正在读取文件..."})
                    df = None
                    try:
                        self.msg_queue.put({'type': 'status', 'message': "优先尝试使用 Polars 快速读取 Excel..."})
                        import polars as pl
                        df_pl = pl.read_excel(
                            file_path,
                            sheet_name=sheet_name,
                            read_options={'header_row': 3}
                        )
                        df = pd.DataFrame(df_pl.to_dict(as_series=False))
                        df.columns = [col.replace('__UNNAMED__', 'Unnamed: ') if str(col).startswith('__UNNAMED__') else col for col in df.columns]
                    except Exception as pl_err:
                        self.msg_queue.put({'type': 'status', 'message': f"Polars 读取 Excel 失败: {str(pl_err)}，切换至 Pandas Calamine 读取..."})
                        df = None
                        
                    if df is None:
                        engine = 'calamine' if HAS_CALAMINE else 'openpyxl'
                        df = pd.read_excel(
                            file_path,
                            sheet_name=sheet_name,
                            skiprows=3,
                            header=0,
                            engine=engine
                        )
                    
                    # 删除第5-12行（对应原始的第8行）
                    df = df.iloc[8:].reset_index(drop=True)
                    
                    # 清理列标题中的括号及其内容，但保留编号
                    def clean_column_name(col):
                        col = str(col)
                        # 如果列名中没有括号，直接返回原格式
                        if '（' not in col and '(' not in col:
                            return col
                        
                        parts = col.split()
                        if len(parts) > 1 and parts[-1].isdigit():
                            # 处理带数字的列名
                            number = parts[-1]
                            prefix = ' '.join(parts[:-1])
                            if '（' in prefix or '(' in prefix:
                                pos1 = prefix.find('（')
                                pos2 = prefix.find('(')
                                if pos1 == -1: pos1 = len(prefix)
                                if pos2 == -1: pos2 = len(prefix)
                                pos = min(pos1, pos2)
                                prefix = prefix[:pos]
                            return f"{prefix.strip()} {number}"
                        else:
                            # 处理不带数字的列名
                            pos1 = col.find('（')
                            pos2 = col.find('(')
                            if pos1 == -1: pos1 = len(col)
                            if pos2 == -1: pos2 = len(col)
                            pos = min(pos1, pos2)
                            return col[:pos].strip()
                    
                    df.columns = [clean_column_name(col) for col in df.columns]
                    
                    # 创建结果DataFrame
                    result_df = pd.DataFrame()
                    
                    # 添加物理时间列
                    result_df['物理时间'] = df.iloc[:, 0]
                    
                    # 添加第一组的第二列
                    result_df[df.columns[1]] = df.iloc[:, 1]
                    
                    # 添加后续每组的中间列
                    num_groups = len(df.columns) // 3
                    for group in range(1, num_groups):
                        col_idx = group * 3 + 1
                        result_df[df.columns[col_idx]] = df.iloc[:, col_idx]
                    
                    # 计算总耗时
                    end_time = time.time()
                    elapsed_time = end_time - start_time
                    self.msg_queue.put({'type': 'status', 'message': f"文件读取完成，耗时: {elapsed_time:.2f}秒"})
                    
                    # 保存处理后的文件
                    csv_path = os.path.splitext(file_path)[0] + '_processed.csv'
                    result_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                    self.msg_queue.put({'type': 'status', 'message': f"数据处理完成并已保存到：{csv_path}"})
                    
                    self.msg_queue.put({'type': 'raw_done', 'df': result_df, 'message': f"数据处理完成并已保存到：{csv_path}"})
                    
                except Exception as e:
                    self.msg_queue.put({'type': 'error', 'message': f"读取Excel文件失败: {str(e)}"})
                    return

            elif file_type == "battery":
                self.msg_queue.put({'type': 'status', 'message': "开始读取电池数据...", 'clear': True})
                try:
                    start_time = time.time()
                    file_ext = os.path.splitext(file_path)[1].lower()
                    
                    try:
                        start_row = int(start_row_str) - 1  # 转换为0基索引
                        if start_row < 0:
                            start_row = 0
                    except ValueError:
                        start_row = 0

                    if file_ext == '.csv':
                        self.msg_queue.put({'type': 'status', 'message': "正在读取CSV文件..."})
                        df = None
                        
                        # 第一步：优先尝试使用 Polars 极速读取（支持 UTF-8 编码且不依赖 pyarrow）
                        try:
                            self.msg_queue.put({'type': 'status', 'message': "尝试使用 Polars 极速读取..."})
                            import polars as pl
                            df_pl = pl.read_csv(file_path, skip_rows=start_row, ignore_errors=True, infer_schema_length=10000)
                            df = pd.DataFrame(df_pl.to_dict(as_series=False))
                            if not df.empty:
                                end_time = time.time()
                                self.msg_queue.put({'type': 'status', 'message': f"使用 Polars 成功读取文件，耗时: {end_time - start_time:.2f}秒"})
                        except ImportError:
                            self.msg_queue.put({'type': 'status', 'message': "未安装 Polars，切换至 Pandas 兼容模式..."})
                        except Exception as pl_err:
                            self.msg_queue.put({'type': 'status', 'message': f"Polars 读取失败（可能编码非 UTF-8 或格式不符）: {str(pl_err)}，切换至 Pandas 兼容模式..."})
                            df = None
                        
                        # 第二步：如果 Polars 失败，自动降级至原有的 Pandas 循环编码读取逻辑
                        if df is None:
                            try:
                                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                    total_lines = sum(1 for _ in f)
                                self.msg_queue.put({'type': 'status', 'message': f"文件总行数: {total_lines}"})
                            except Exception:
                                total_lines = None
                            
                            encodings = ['utf-8', 'gb18030', 'gbk', 'gb2312', 'ascii']
                            
                            for encoding in encodings:
                                try:
                                    self.msg_queue.put({'type': 'status', 'message': f"尝试使用 {encoding} 编码读取..."})
                                    chunks = []
                                    for i, chunk in enumerate(pd.read_csv(file_path, encoding=encoding, 
                                                                         chunksize=self.CHUNK_SIZE, 
                                                                         skiprows=start_row if start_row > 0 else None)):
                                        chunks.append(chunk)
                                        if total_lines:
                                            current_rows = min((i + 1) * self.CHUNK_SIZE, total_lines)
                                            progress = min(100, current_rows / total_lines * 100)
                                            self.msg_queue.put({'type': 'status', 'message': f"读取进度: {progress:.1f}% ({current_rows}/{total_lines}行)"})
                                    
                                    df = pd.concat(chunks, ignore_index=True)
                                    if not df.empty:
                                        end_time = time.time()
                                        self.msg_queue.put({'type': 'status', 'message': f"使用 {encoding} 编码成功读取文件，耗时: {end_time - start_time:.2f}秒"})
                                        break
                                except Exception as e:
                                    self.msg_queue.put({'type': 'status', 'message': f"使用 {encoding} 编码读取失败: {str(e)}"})
                                    continue
                                    
                        if df is None:
                            self.msg_queue.put({'type': 'error', 'message': "读取电池CSV文件失败，所有编码尝试均已失败。"})
                            return
                        df = df.dropna(how='all')
                    else:
                        self.msg_queue.put({'type': 'status', 'message': "正在读取文件..."})
                        df = None
                        try:
                            self.msg_queue.put({'type': 'status', 'message': "优先尝试使用 Polars 快速读取 Excel..."})
                            import polars as pl
                            read_opts = {'header_row': start_row} if start_row > 0 else {}
                            df_pl = pl.read_excel(
                                file_path,
                                sheet_name=sheet_name,
                                read_options=read_opts
                            )
                            df = pd.DataFrame(df_pl.to_dict(as_series=False))
                            df.columns = [col.replace('__UNNAMED__', 'Unnamed: ') if str(col).startswith('__UNNAMED__') else col for col in df.columns]
                            df = df.dropna(how='all')
                        except Exception as pl_err:
                            self.msg_queue.put({'type': 'status', 'message': f"Polars 读取 Excel 失败: {str(pl_err)}，切换至 Pandas Calamine 读取..."})
                            df = None

                        if df is None:
                            use_calamine = HAS_CALAMINE
                            if use_calamine:
                                self.msg_queue.put({'type': 'status', 'message': "正在使用 Calamine 快速读取..."})
                                try:
                                    df = pd.read_excel(
                                        file_path,
                                        sheet_name=sheet_name,
                                        skiprows=start_row if start_row > 0 else None,
                                        engine='calamine'
                                    )
                                    df = df.dropna(how='all')
                                except Exception as e:
                                    self.msg_queue.put({'type': 'status', 'message': f"Calamine 读取失败, 尝试标准方法: {str(e)}"})
                                    use_calamine = False
                            
                            if not use_calamine:
                                self.msg_queue.put({'type': 'status', 'message': "正在使用 openpyxl read-only 模式读取..."})
                                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                                sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                                total_rows = sheet.max_row
                                self.msg_queue.put({'type': 'status', 'message': f"文件总行数: {total_rows or '未知'}"})
                                
                                data = []
                                header = None
                                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                                    if r_idx < start_row:
                                        continue
                                    if r_idx == start_row:
                                        header = list(row)
                                    else:
                                        data.append(row)
                                    
                                    if r_idx > start_row and (r_idx - start_row) % 20000 == 0:
                                        percent = ((r_idx - start_row) / total_rows * 100) if total_rows else 0
                                        self.msg_queue.put({'type': 'status', 'message': f"读取进度: {percent:.1f}% ({r_idx - start_row}/{total_rows - start_row if total_rows else '未知'}行)"})
                                
                                df = pd.DataFrame(data, columns=header)
                                df = df.dropna(how='all')
                                wb.close()
                    # Auto-detect default columns
                    cycle_default = None
                    step_default = None
                    time_default = None
                    
                    for col in df.columns:
                        col_lower = str(col).lower()
                        if cycle_default is None and any(x in col_lower for x in ['循环序号', '循环号', '循环', 'cycle']):
                            cycle_default = col
                        if step_default is None and any(x in col_lower for x in ['工步状态', '工步', '工作状态', 'step', 'state', 'status']):
                            step_default = col
                        if time_default is None and any(x in col_lower for x in ['时间', 'time']):
                            if not col_lower.endswith('_时间差(s)'):
                                time_default = col

                    if not cycle_default and len(df.columns) > 0:
                        cycle_default = df.columns[0]
                    if not step_default and len(df.columns) > 0:
                        step_default = df.columns[min(1, len(df.columns)-1)]
                    if not time_default and len(df.columns) > 0:
                        time_default = df.columns[min(2, len(df.columns)-1)]

                    # Convert datetime columns to time difference in seconds
                    import pandas.api.types as ptypes
                    import warnings
                    self.msg_queue.put({'type': 'status', 'message': "正在转换默认时间列并计算时间差(s)..."})
                    
                    if time_default and time_default in df.columns:
                        col = time_default
                        new_col_name = f"{col}_时间差(s)"
                        if ptypes.is_datetime64_any_dtype(df[col]):
                            first_valid = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
                            if first_valid is not None:
                                df[new_col_name] = (df[col] - first_valid).dt.total_seconds()
                        elif df[col].dtype == 'object' or ptypes.is_string_dtype(df[col]):
                            non_nulls = df[col].dropna()
                            if not non_nulls.empty:
                                try:
                                    with warnings.catch_warnings():
                                        warnings.filterwarnings("ignore", category=UserWarning, message=".*Could not infer format.*")
                                        pd.to_datetime(non_nulls.head(10), errors='raise')
                                        times = pd.to_datetime(df[col], errors='coerce')
                                    first_valid = times.dropna().iloc[0] if not times.dropna().empty else None
                                    if first_valid is not None:
                                        df[new_col_name] = (times - first_valid).dt.total_seconds()
                                except Exception:
                                    pass

                    # Populate filter values in background thread
                    self.msg_queue.put({'type': 'status', 'message': "正在生成工步和循环筛选项..."})
                    cycles = ['全部']
                    if cycle_default:
                        try:
                            unique_vals = df[cycle_default].dropna().unique()
                            def safe_sort_key(val):
                                try:
                                    return (0, float(val))
                                except (ValueError, TypeError):
                                    return (1, str(val))
                            unique_vals = sorted(unique_vals, key=safe_sort_key)
                            for x in unique_vals:
                                try:
                                    if float(x).is_integer():
                                        cycles.append(str(int(x)))
                                    else:
                                        cycles.append(str(x))
                                except ValueError:
                                    cycles.append(str(x))
                        except Exception:
                            cycles = ['全部'] + [str(x) for x in df[cycle_default].dropna().unique()]
                            
                    steps = ['全部']
                    if step_default:
                        try:
                            steps = ['全部'] + [str(x) for x in df[step_default].dropna().unique()]
                        except Exception:
                            pass

                    end_time = time.time()
                    self.msg_queue.put({
                        'type': 'done', 
                        'df': df,
                        'cycles': cycles,
                        'steps': steps,
                        'cycle_default': cycle_default,
                        'step_default': step_default,
                        'time_default': time_default,
                        'message': f"文件读取完成，耗时: {end_time - start_time:.2f}秒，共读取 {len(df)} 行数据"
                    })
                    
                except Exception as e:
                    self.msg_queue.put({'type': 'error', 'message': f"读取电池数据失败: {str(e)}"})
                    return
            
            else:
                # 获取起始行
                try:
                    start_row = int(start_row_str) - 1  # 转换为0基索引
                    if start_row < 0:
                        start_row = 0
                except ValueError:
                    start_row = 0
                
                # 根据文件类型读取数据
                file_ext = os.path.splitext(file_path)[1].lower()
                if file_ext == '.csv':
                    self.msg_queue.put({'type': 'status', 'message': "开始读取CSV文件...", 'clear': True})
                    start_time = time.time()
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            total_lines = sum(1 for _ in f)
                        self.msg_queue.put({'type': 'status', 'message': f"文件总行数: {total_lines}"})
                    except Exception:
                        total_lines = None
                    
                    encodings = ['utf-8', 'gb18030', 'gbk', 'gb2312', 'ascii']
                    result_df = None
                    
                    for encoding in encodings:
                        try:
                            self.msg_queue.put({'type': 'status', 'message': f"尝试使用 {encoding} 编码读取..."})
                            chunks = []
                            for i, chunk in enumerate(pd.read_csv(file_path, encoding=encoding, 
                                                                chunksize=self.CHUNK_SIZE, 
                                                                skiprows=start_row if start_row > 0 else None)):
                                chunks.append(chunk)
                                if total_lines:
                                    current_rows = min((i + 1) * self.CHUNK_SIZE, total_lines)
                                    progress = min(100, current_rows / total_lines * 100)
                                    self.msg_queue.put({'type': 'status', 'message': f"读取进度: {progress:.1f}% ({current_rows}/{total_lines}行)"})
                            
                            result_df = pd.concat(chunks, ignore_index=True)
                            if not result_df.empty:
                                end_time = time.time()
                                self.msg_queue.put({'type': 'status', 'message': f"使用 {encoding} 编码成功读取文件，耗时: {end_time - start_time:.2f}秒"})
                                break
                        except Exception as e:
                            self.msg_queue.put({'type': 'status', 'message': f"使用 {encoding} 编码读取失败: {str(e)}"})
                            continue
                    
                    if result_df is not None:
                        self.msg_queue.put({'type': 'done', 'df': result_df, 'message': f"CSV文件读取完成，共读取 {len(result_df)} 行数据"})
                    else:
                        self.msg_queue.put({'type': 'error', 'message': "读取CSV文件失败，所有编码尝试均已失败。"})
                            
                elif file_ext in ['.xlsx', '.xls']:
                    self.msg_queue.put({'type': 'status', 'message': "开始读取Excel文件...", 'clear': True})
                    try:
                        start_time = time.time()
                        self.msg_queue.put({'type': 'status', 'message': "正在计算文件大小..."})
                        try:
                            wb = openpyxl.load_workbook(file_path, read_only=True)
                            if sheet_name in wb.sheetnames:
                                sheet = wb[sheet_name]
                                total_rows = sheet.max_row
                            else:
                                total_rows = wb.active.max_row
                            wb.close()
                        except Exception:
                            total_rows = None
                        
                        if total_rows:
                            self.msg_queue.put({'type': 'status', 'message': f"文件总行数: {total_rows}"})
                            if total_rows <= start_row:
                                self.msg_queue.put({'type': 'error', 'message': "错误：起始行超出文件总行数"})
                                return
                        
                        result_df = None
                        try:
                            self.msg_queue.put({'type': 'status', 'message': "优先尝试使用 Polars 快速读取 Excel..."})
                            import polars as pl
                            read_opts = {'header_row': start_row} if start_row > 0 else {}
                            df_pl = pl.read_excel(
                                file_path,
                                sheet_name=sheet_name,
                                read_options=read_opts
                            )
                            result_df = pd.DataFrame(df_pl.to_dict(as_series=False))
                            result_df.columns = [col.replace('__UNNAMED__', 'Unnamed: ') if str(col).startswith('__UNNAMED__') else col for col in result_df.columns]
                        except Exception as pl_err:
                            self.msg_queue.put({'type': 'status', 'message': f"Polars 读取 Excel 失败: {str(pl_err)}，切换至 Pandas Calamine 读取..."})
                            result_df = None

                        if result_df is None:
                            use_calamine = HAS_CALAMINE
                            if use_calamine:
                                self.msg_queue.put({'type': 'status', 'message': "正在使用 Calamine 快速读取..."})
                                try:
                                    result_df = pd.read_excel(
                                        file_path,
                                        sheet_name=sheet_name,
                                        skiprows=start_row,
                                        engine='calamine'
                                    )
                                except Exception as e:
                                    self.msg_queue.put({'type': 'status', 'message': f"Calamine 读取失败, 尝试标准方法: {str(e)}"})
                                    use_calamine = False
                                    
                            if not use_calamine:
                                self.msg_queue.put({'type': 'status', 'message': "正在使用 openpyxl read-only 模式读取..."})
                                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                                sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                                
                                data = []
                                header = None
                                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                                    if r_idx < start_row:
                                        continue
                                    if r_idx == start_row:
                                        header = list(row)
                                    else:
                                        data.append(row)
                                    
                                    if r_idx > start_row and (r_idx - start_row) % 20000 == 0:
                                        percent = ((r_idx - start_row) / total_rows * 100) if total_rows else 0
                                        self.msg_queue.put({'type': 'status', 'message': f"读取进度: {percent:.1f}% ({r_idx - start_row}/{total_rows - start_row if total_rows else '未知'}行)"})
                                
                                result_df = pd.DataFrame(data, columns=header)
                                wb.close()
                        end_time = time.time()
                        self.msg_queue.put({'type': 'done', 'df': result_df, 'message': f"文件读取完成，耗时: {end_time - start_time:.2f}秒，共读取 {len(result_df)} 行数据"})
                        
                    except Exception as e:
                        self.msg_queue.put({'type': 'error', 'message': f"读取Excel文件失败: {str(e)}"})
                        return
                else:
                    self.msg_queue.put({'type': 'error', 'message': "错误：不支持的文件格式"})
                    return
        except Exception as e:
            self.msg_queue.put({'type': 'error', 'message': f"后台处理发生未知错误: {str(e)}"})

    def on_selection_change(self, event):
        """当任何Listbox的选择改变时更新状态"""
        # 只更新发生改变的Listbox的选择状态
        widget = event.widget
        for i, listbox in enumerate(self.y_listboxes):
            if listbox == widget:  # 只更新发生改变的Listbox
                self.y_selections[i] = [listbox.get(idx) for idx in listbox.curselection()]
                break
        
        # 延迟更新图表以防连击/拖动多选卡死
        self.delayed_update()

    def on_cycle_col_changed(self, event=None):
        if self.result_df is not None:
            cycle_col_name = self.cycle_col.get()
            if cycle_col_name in self.result_df.columns:
                try:
                    unique_vals = self.result_df[cycle_col_name].dropna().unique()
                    def safe_sort_key(val):
                        try:
                            return (0, float(val))
                        except (ValueError, TypeError):
                            return (1, str(val))
                    unique_vals = sorted(unique_vals, key=safe_sort_key)
                    cycles = ['全部']
                    for x in unique_vals:
                        try:
                            if float(x).is_integer():
                                cycles.append(str(int(x)))
                            else:
                                cycles.append(str(x))
                        except ValueError:
                            cycles.append(str(x))
                except Exception:
                    cycles = ['全部'] + [str(x) for x in self.result_df[cycle_col_name].dropna().unique()]
                self.cycle_filter_combo['values'] = cycles
                self.cycle_filter_combo.set('全部')
                self.delayed_update()

    def on_step_col_changed(self, event=None):
        if self.result_df is not None:
            step_col_name = self.step_col.get()
            if step_col_name in self.result_df.columns:
                steps = ['全部'] + [str(x) for x in self.result_df[step_col_name].dropna().unique()]
                self.step_filter_combo['values'] = steps
                self.step_filter_combo.set('全部')
                self.delayed_update()

    def on_time_col_changed(self, event=None):
        if self.result_df is not None:
            time_col_name = self.time_col.get()
            if time_col_name in self.result_df.columns:
                new_col_name = f"{time_col_name}_时间差(s)"
                if new_col_name not in self.result_df.columns:
                    self.set_buttons_state(False)
                    self.update_status(f"正在后台计算 {time_col_name} 的时间差(s)...")
                    threading.Thread(target=self._bg_calc_time_diff, args=(time_col_name, new_col_name), daemon=True).start()
                    return
            self.update_listboxes()
            self.delayed_update()

    def _bg_calc_time_diff(self, time_col_name, new_col_name):
        try:
            import pandas.api.types as ptypes
            col_data = self.result_df[time_col_name]
            
            # 1. 优先尝试作为数值时间列处理（已经是数字，或者可以无损转为数字）
            numeric_col = pd.to_numeric(col_data, errors='coerce')
            non_null_count = col_data.dropna().shape[0]
            
            if non_null_count > 0 and numeric_col.notna().sum() >= 0.95 * non_null_count:
                non_nulls = numeric_col.dropna()
                first_valid = non_nulls.iloc[0] if not non_nulls.empty else None
                if first_valid is not None:
                    # 对于已经是数值的列（如秒数 0,1,2,3 ），直接做差值，使其从 0 开始
                    self.result_df[new_col_name] = numeric_col - first_valid
            # 2. 如果是 Datetime64 格式
            elif ptypes.is_datetime64_any_dtype(col_data):
                non_nulls = col_data.dropna()
                first_valid = non_nulls.iloc[0] if not non_nulls.empty else None
                if first_valid is not None:
                    self.result_df[new_col_name] = (col_data - first_valid).dt.total_seconds()
            # 3. 尝试解析文本为 Datetime 格式
            else:
                non_nulls = col_data.dropna()
                if not non_nulls.empty:
                    try:
                        pd.to_datetime(non_nulls.head(10), errors='raise')
                        times = pd.to_datetime(col_data, errors='coerce')
                        valid_times = times.dropna()
                        first_valid = valid_times.iloc[0] if not valid_times.empty else None
                        if first_valid is not None:
                            self.result_df[new_col_name] = (times - first_valid).dt.total_seconds()
                    except Exception:
                        pass
                        
            self.msg_queue.put({
                'type': 'time_diff_done',
                'message': f"计算 {time_col_name} 时间差(s)完成"
            })
        except Exception as e:
            self.msg_queue.put({
                'type': 'error',
                'message': f"计算时间差失败: {str(e)}"
            })

    def update_font_and_plot(self):
        """更新字体设置并重新绘制图例"""
        try:
            # 更新matplotlib的字体设置
            plt.rcParams['font.family'] = self.font_family.get()
            font_size = self.safe_float_convert(self.font_size.get(), 15)
            plt.rcParams['font.size'] = font_size
            plt.rcParams['axes.titlesize'] = font_size
            plt.rcParams['axes.labelsize'] = font_size
            plt.rcParams['xtick.labelsize'] = font_size
            plt.rcParams['ytick.labelsize'] = font_size
            plt.rcParams['legend.fontsize'] = font_size
            
            # 只更新图例
            self.update_legend_only()
            
        except Exception as e:
            return

    def update_plot(self):
        """添加更新频率限制"""
        if getattr(self, '_is_loading_settings', False):
            return
        current_time = time.time()
        if current_time - self._last_plot_time < 0.1:  # 限制更新频率为10Hz
            return
        self._last_plot_time = current_time
        self.plot_data()

    def plot_y_axis(self, axis_index):
        """根据点击的按钮绘制对应的Y轴数据"""
        # 保存当前选择状态
        widget = self.y_listboxes[axis_index]
        self.y_selections[axis_index] = [widget.get(idx) for idx in widget.curselection()]
        
        # 检查是否有选择的数据
        if not self.y_selections[axis_index]:
            return
        
        # 绘制图表
        self.plot_data(clicked_axis=axis_index)

    def plot_data(self, clicked_axis=None):
        try:
            if self.result_df is None:
                return
            
            df_to_plot = self.result_df
            
            if self.file_type.get() == "battery":
                cycle_col_name = self.cycle_col.get()
                step_col_name = self.step_col.get()
                cycle_val = self.cycle_filter.get()
                step_val = self.step_filter.get()
                
                if cycle_col_name and cycle_col_name in df_to_plot.columns:
                    if cycle_val != "全部" and cycle_val != "":
                        try:
                            try:
                                target_val = int(cycle_val)
                                df_to_plot = df_to_plot[df_to_plot[cycle_col_name] == target_val]
                            except ValueError:
                                try:
                                    target_val = float(cycle_val)
                                    df_to_plot = df_to_plot[df_to_plot[cycle_col_name] == target_val]
                                except ValueError:
                                    df_to_plot = df_to_plot[df_to_plot[cycle_col_name].astype(str) == str(cycle_val)]
                        except Exception:
                            pass
                if step_col_name and step_col_name in df_to_plot.columns:
                    if step_val != "全部" and step_val != "":
                        try:
                            df_to_plot = df_to_plot[df_to_plot[step_col_name] == step_val]
                        except Exception:
                            try:
                                df_to_plot = df_to_plot[df_to_plot[step_col_name].astype(str) == str(step_val)]
                            except Exception:
                                pass
            
            if df_to_plot.empty:
                self.ax.clear()
                self.canvas.draw()
                return

            # 数据自动降采样逻辑，以提升绘图效率，防止大数量级下界面卡死
            if self.auto_downsample.get():
                num_rows = len(df_to_plot)
                max_pts = self.max_plot_points.get()
                if max_pts != "无限制":
                    try:
                        max_pts_val = int(max_pts)
                        if num_rows > max_pts_val:
                            step = num_rows // max_pts_val
                            if step > 1:
                                df_to_plot = df_to_plot.iloc[::step]
                                self.update_status(f"提示：绘图数据已自动降采样（从 {num_rows} 行降至 {len(df_to_plot)} 行），以提升显示效率")
                    except Exception as e:
                        self.logger.error(f"数据降采样失败: {str(e)}")
                
            def set_axis_style(ax):
                ax.tick_params(axis='both', direction='in', width=float(self.frame_width.get()), length=6, 
                             labelsize=int(self.font_size.get()), color='black')
                for spine in ax.spines.values():
                    spine.set_linewidth(float(self.frame_width.get()))
                    spine.set_color('black')
                    
            y1_data = self.y_selections[0]
            y2_data = self.y_selections[1]
            y3_data = self.y_selections[2]
            
            if not any([y1_data, y2_data, y3_data]):
                return
                
            if clicked_axis == 0 or clicked_axis is None:
                self.fig.clf()
                
                # 恢复当前 canvas 控件实际大小对应的 figure 尺寸，防止 clf() 重置 figure 尺寸到默认值 6.4x4.8
                canvas_widget = self.canvas.get_tk_widget()
                w_px = canvas_widget.winfo_width()
                h_px = canvas_widget.winfo_height()
                if w_px > 1 and h_px > 1:
                    self.fig.set_size_inches(w_px / self.fig.dpi, h_px / self.fig.dpi, forward=False)
                    
                self.ax = self.fig.add_subplot(111)
                if clicked_axis == 0:
                    y2_data = []
                    y3_data = []
            elif clicked_axis == 1:
                if not y1_data:
                    return
                y3_data = []
            elif clicked_axis == 2:
                if not y1_data or not y2_data:
                    return
                

            
            x_col = self.x_axis.get()
            font_size = int(self.font_size.get())
            font_family = self.font_family.get()
            
            # Safe X categorical check
            if x_col and x_col in df_to_plot.columns:
                import pandas.api.types as ptypes
                if df_to_plot[x_col].dtype == 'object' or ptypes.is_string_dtype(df_to_plot[x_col]):
                    unique_count = df_to_plot[x_col].nunique()
                    if unique_count > 1000:
                        messagebox.showwarning("警告", f"X轴 '{x_col}' 包含大量文本值 ({unique_count}个唯一值)，直接绘制会导致界面卡死。\n请选择时间差（例如包含'时间差(s)'的列）等数值列作为X轴。")
                        return

            all_lines = []
            all_labels = []
            
            if y1_data:
                set_axis_style(self.ax)
                color_map = self.color_schemes_dict[self.color_schemes[0].get()]
                for i, col in enumerate(y1_data):
                    color = plt.cm.tab10(i % 10) if color_map is None else color_map(i % color_map.N)
                    line = self.ax.plot(df_to_plot[x_col], df_to_plot[col],
                                      label=col, 
                                      linewidth=self.safe_float_convert(self.line_width.get(), 1.5),
                                      color=color,
                                      linestyle=self.line_styles_dict[self.line_styles[0].get()])
                    all_lines.extend(line)
                    all_labels.append(col)
                self.ax.set_ylabel(self.y_settings[0]['title'].get(), fontsize=font_size, fontfamily=font_family, color='black', labelpad=10)
                try:
                    ymin = float(self.y_settings[0]['min'].get())
                    ymax = float(self.y_settings[0]['max'].get())
                    if ymin < ymax:
                        self.ax.set_ylim(ymin, ymax)
                except Exception:
                    pass
                
            if y2_data:
                ax2 = self.ax.twinx()
                ax2.spines['right'].set_position(('outward', 0))
                set_axis_style(ax2)
                color_map = self.color_schemes_dict[self.color_schemes[1].get()]
                y2_lines_temp = []
                y2_labels_temp = []
                for i, col in enumerate(y2_data):
                    color = plt.cm.tab10(i % 10) if color_map is None else color_map(i % color_map.N)
                    line = ax2.plot(df_to_plot[x_col], df_to_plot[col],
                                      label=col, 
                                      linewidth=self.safe_float_convert(self.line_width.get(), 1.5),
                                      color=color,
                                      linestyle=self.line_styles_dict[self.line_styles[1].get()])
                    y2_lines_temp.extend(line)
                    y2_labels_temp.append(col)
                all_lines.extend(y2_lines_temp)
                all_labels.extend(y2_labels_temp)
                ax2.set_ylabel(self.y_settings[1]['title'].get(), fontsize=font_size, fontfamily=font_family, color='black', labelpad=15)
                try:
                    ymin = float(self.y_settings[1]['min'].get())
                    ymax = float(self.y_settings[1]['max'].get())
                    if ymin < ymax:
                        ax2.set_ylim(ymin, ymax)
                except Exception:
                    pass
                
            if y3_data:
                ax3 = self.ax.twinx()
                ax3.spines['right'].set_position(('outward', 70))
                set_axis_style(ax3)
                color_map = self.color_schemes_dict[self.color_schemes[2].get()]
                for i, col in enumerate(y3_data):
                    color = plt.cm.tab10(i % 10) if color_map is None else color_map(i % color_map.N)
                    line = ax3.plot(df_to_plot[x_col], df_to_plot[col],
                                      label=col, 
                                      linewidth=self.safe_float_convert(self.line_width.get(), 1.5),
                                      color=color,
                                      linestyle=self.line_styles_dict[self.line_styles[2].get()])
                    all_lines.extend(line)
                    all_labels.append(col)
                ax3.set_ylabel(self.y_settings[2]['title'].get(), fontsize=font_size, fontfamily=font_family, color='black', labelpad=15)
                try:
                    ymin = float(self.y_settings[2]['min'].get())
                    ymax = float(self.y_settings[2]['max'].get())
                    if ymin < ymax:
                        ax3.set_ylim(ymin, ymax)
                except Exception:
                    pass
                
            if all_lines and self.legend_visible.get():
                y1_end = len(y1_data) if y1_data else 0
                y2_end = y1_end + (len(y2_data) if y2_data else 0)
                y1_lines = all_lines[:y1_end] if y1_data else []
                y2_lines = all_lines[y1_end:y2_end] if y2_data else []
                y3_lines = all_lines[y2_end:] if y3_data else []
                y1_labels = all_labels[:y1_end] if y1_data else []
                y2_labels = all_labels[y1_end:y2_end] if y2_data else []
                y3_labels = all_labels[y2_end:] if y3_data else []
                
                legend_y = self.safe_float_convert(self.legend_y.get(), 1.02)
                positions = self.get_legend_positions()
                
                for ax in self.fig.axes:
                    if ax.get_legend() is not None:
                        ax.get_legend().remove()
                    
                if y1_data:
                    leg1 = self.ax.legend(y1_lines, y1_labels, loc='upper left', bbox_to_anchor=(positions[0], legend_y), ncol=1, frameon=False, fontsize=font_size)
                    self.ax.add_artist(leg1)
                if y2_data:
                    leg2 = self.ax.legend(y2_lines, y2_labels, loc='upper left', bbox_to_anchor=(positions[1], legend_y), ncol=1, frameon=False, fontsize=font_size)
                    self.ax.add_artist(leg2)
                if y3_data:
                    leg3 = self.ax.legend(y3_lines, y3_labels, loc='upper left', bbox_to_anchor=(positions[2], legend_y), ncol=1, frameon=False, fontsize=font_size)
                    self.ax.add_artist(leg3)
                    
            try:
                self.ax.ticklabel_format(axis='x', style='sci', scilimits=(-3, 6))
                self.ax.xaxis.get_offset_text().set_fontsize(font_size)
                self.ax.xaxis.get_offset_text().set_fontfamily(font_family)
            except Exception:
                pass
                
            # 动态计算最合理的边界占比，保证刻度不被裁剪且绘图区域最大化
            right_margin, left_margin = self.get_dynamic_margins(y1_data, y2_data, y3_data)
                
            self.fig.subplots_adjust(right=right_margin, left=left_margin, top=0.90, bottom=0.12)
            self.canvas.draw()
            
        except Exception as e:
            messagebox.showerror("错误", f"绘图失败: {str(e)}")

    def get_dynamic_margins(self, y1_data, y2_data, y3_data):
        """根据当前图纸的实际像素宽度和字体大小，动态计算并返回左右边距百分比"""
        try:
            # 优先使用 Tkinter Canvas 控件的实际物理像素宽度，防止 clf() 重置 figure 尺寸导致的计算偏差
            canvas_widget = self.canvas.get_tk_widget()
            fig_width_px = canvas_widget.winfo_width()
            
            # 如果控件尚未渲染，使用基于 figure 的理论尺寸
            if fig_width_px <= 1:
                fig_width_px = self.fig.get_figwidth() * self.fig.dpi
            if fig_width_px <= 1:
                fig_width_px = 1200 # 容错基准值
                
            font_size = int(self.font_size.get())
            left_margin_px = max(90, int(font_size * 6.5))
            
            if y3_data:
                right_margin_px = max(200, int(font_size * 13.5))
                max_right_percent = 0.85  # 预留至少 12% 的空间给右侧双 Y 轴 (Y2, Y3)
            elif y2_data:
                right_margin_px = max(100, int(font_size * 7.0))
                max_right_percent = 0.93  # 预留至少 7% 的空间给右侧单 Y 轴 (Y2)
            else:
                right_margin_px = max(45, int(font_size * 3.0))
                max_right_percent = 0.96  # 预留至少 4% 的空间给无右侧轴的情况
                
            left_margin = max(0.08, min(left_margin_px / fig_width_px, 0.2))
            right_margin = max(0.4, min(1.0 - (right_margin_px / fig_width_px), max_right_percent))
            return right_margin, left_margin
        except Exception:
            if y3_data:
                return 0.84, 0.08
            elif y2_data:
                return 0.90, 0.08
            else:
                return 0.95, 0.08

    def on_window_resize(self, event):
        # 只在根窗口大小改变时更新图表大小
        if event.widget != self.root:
            return
        if hasattr(self, 'fig'):
            # 动态更新 figure 尺寸以匹配实际控件大小，防止高分辨率下部分区域被截断
            canvas_widget = self.canvas.get_tk_widget()
            w_px = canvas_widget.winfo_width()
            h_px = canvas_widget.winfo_height()
            if w_px > 1 and h_px > 1:
                self.fig.set_size_inches(w_px / self.fig.dpi, h_px / self.fig.dpi, forward=False)
                
            # 根据当前显示的Y轴数量动态调整边距
            y1_data = self.y_selections[0]
            y2_data = self.y_selections[1]
            y3_data = self.y_selections[2]
            
            right_margin, left_margin = self.get_dynamic_margins(y1_data, y2_data, y3_data)
                
            self.fig.subplots_adjust(
                right=right_margin, 
                left=left_margin, 
                top=0.90,
                bottom=0.12,
                wspace=0.2
            )
            # 使用 draw_idle 替代 draw 提升重绘性能并防止 configure 死锁
            self.canvas.draw_idle()

    def update_legend_only(self):
        """只更新图例位置，不重新绘制整个图表（带防抖）"""
        if getattr(self, '_is_loading_settings', False):
            return
        if not hasattr(self, 'ax') or not self.legend_visible.get():
            return
            
        if hasattr(self, '_legend_timer'):
            try:
                self.root.after_cancel(self._legend_timer)
            except Exception:
                pass
                
        def do_update():
            try:
                # 获取当前的图例位置设置
                legend_y = self.safe_float_convert(self.legend_y.get(), 1.02)
                font_size = self.safe_float_convert(self.font_size.get(), 15)
                
                # 获取当前选择的数据
                y1_data = self.y_selections[0]
                y2_data = self.y_selections[1]
                y3_data = self.y_selections[2]
                
                # 获取所有轴对象
                axes = [self.ax]
                for axis in self.ax.figure.axes:
                    if axis != self.ax:
                        axes.append(axis)
                
                # 移除其他轴的图例
                for ax in axes[1:]:
                    if ax.get_legend() is not None:
                        ax.get_legend().remove()
                
                # 获取所有线条和标签
                all_lines = []
                all_labels = []
                
                # 从每个轴获取可见的线条和标签
                for ax in axes:
                    for line in ax.get_lines():
                        if line.get_visible():
                            all_lines.append(line)
                            all_labels.append(line.get_label())
                
                # 计算每个Y轴的线条数量
                y1_end = len(y1_data) if y1_data else 0
                y2_end = y1_end + (len(y2_data) if y2_data else 0)
                
                # 分离每个Y轴的线条和标签
                y1_lines = all_lines[:y1_end] if y1_data else []
                y2_lines = all_lines[y1_end:y2_end] if y2_data else []
                y3_lines = all_lines[y2_end:] if y3_data else []
                
                y1_labels = all_labels[:y1_end] if y1_data else []
                y2_labels = all_labels[y1_end:y2_end] if y2_data else []
                y3_labels = all_labels[y2_end:] if y3_data else []
                
                # 重新创建图例
                if y1_lines:
                    leg1 = self.ax.legend(y1_lines, y1_labels,
                                        loc='upper left',
                                        bbox_to_anchor=(float(self.legend_x_positions_str.get().split(',')[0]), legend_y),
                                        ncol=1,
                                        frameon=False,
                                        fontsize=font_size)
                    self.ax.add_artist(leg1)
                
                if y2_lines:
                    leg2 = self.ax.legend(y2_lines, y2_labels,
                                        loc='upper left',
                                        bbox_to_anchor=(float(self.legend_x_positions_str.get().split(',')[1]), legend_y),
                                        ncol=1,
                                        frameon=False,
                                        fontsize=font_size)
                    self.ax.add_artist(leg2)
                
                if y3_lines:
                    leg3 = self.ax.legend(y3_lines, y3_labels,
                                        loc='upper left',
                                        bbox_to_anchor=(float(self.legend_x_positions_str.get().split(',')[2]), legend_y),
                                        ncol=1,
                                        frameon=False,
                                        fontsize=font_size)
                    self.ax.add_artist(leg3)
                
                # 重新绘制画布
                self.canvas.draw()
                
            except Exception as e:
                self.logger.error(f"更新图例失败: {str(e)}")

        self._legend_timer = self.root.after(300, do_update)

    def clear_all_selections(self):
        """清除所有选择"""
        for i, listbox in enumerate(self.y_listboxes):
            listbox.selection_clear(0, tk.END)
            self.y_selections[i] = []
        # 清除所有图形
        if hasattr(self, 'fig'):
            self.fig.clf()
            self.ax = self.fig.add_subplot(111)
            self.canvas.draw()

    def select_all_y(self):
        """全部选择Y1"""
        if not self.x_axis.get():
            return
        
        self.clear_all_selections()
        x_col = self.x_axis.get()
        
        # 只在Y1中选择所有数据
        listbox = self.y_listboxes[0]
        for i in range(listbox.size()):
            if listbox.get(i) != x_col:
                listbox.selection_set(i)
        # 更新Y1的选择状态
        self.y_selections[0] = [listbox.get(i) for i in listbox.curselection()]
        self.update_plot()

    def safe_float_convert(self, value, default=0.0):
        """安全地转换浮点数"""
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def validate_data(self):
        """验证数据有效性"""
        if not hasattr(self, 'result_df') or self.result_df is None:
            return False
        if not any(self.y_selections):
            return False
        return True

    def clear_memory(self):
        """清理内存"""
        try:
            if hasattr(self, 'result_df'):
                del self.result_df
            if hasattr(self, 'fig'):
                plt.close('all')
            gc.collect()
        except Exception as e:
            print(f"清理内存时出错: {str(e)}")

    def __del__(self):
        """析构函数"""
        self.clear_memory()

    def save_plot_data(self):
        """将当前绘图数据保存为 xlsx 文件"""
        try:
            if self.result_df is None:
                messagebox.showwarning("警告", "当前无有效数据，请先载入并计算数据！")
                return
            
            x_col = self.x_axis.get()
            if not x_col:
                messagebox.showwarning("警告", "请先选择 X 轴数据！")
                return
                
            y1_cols = self.y_selections[0]
            y2_cols = self.y_selections[1]
            y3_cols = self.y_selections[2]
            all_y_cols = y1_cols + y2_cols + y3_cols
            
            if not all_y_cols:
                messagebox.showwarning("警告", "请先选择至少一个 Y 轴进行绘图！")
                return
                
            # 获取当前筛选后的数据子集 (与 plot_data 过滤逻辑一致)
            df_to_plot = self.result_df
            
            if self.file_type.get() == "battery":
                cycle_col_name = self.cycle_col.get()
                step_col_name = self.step_col.get()
                cycle_val = self.cycle_filter.get()
                step_val = self.step_filter.get()
                
                if cycle_col_name and cycle_col_name in df_to_plot.columns:
                    if cycle_val != "全部" and cycle_val != "":
                        try:
                            try:
                                target_val = int(cycle_val)
                                df_to_plot = df_to_plot[df_to_plot[cycle_col_name] == target_val]
                            except ValueError:
                                try:
                                    target_val = float(cycle_val)
                                    df_to_plot = df_to_plot[df_to_plot[cycle_col_name] == target_val]
                                except ValueError:
                                    df_to_plot = df_to_plot[df_to_plot[cycle_col_name].astype(str) == str(cycle_val)]
                        except Exception:
                            pass
                if step_col_name and step_col_name in df_to_plot.columns:
                    if step_val != "全部" and step_val != "":
                        try:
                            df_to_plot = df_to_plot[df_to_plot[step_col_name] == step_val]
                        except Exception:
                            try:
                                df_to_plot = df_to_plot[df_to_plot[step_col_name].astype(str) == str(step_val)]
                            except Exception:
                                pass
            
            if df_to_plot.empty:
                messagebox.showwarning("警告", "筛选后的绘图数据为空，无法保存！")
                return
                
            # 提取选中的 X 轴和 Y 轴列
            selected_cols = []
            if x_col and x_col in df_to_plot.columns:
                selected_cols.append(x_col)
            for col in all_y_cols:
                if col and col in df_to_plot.columns and col not in selected_cols:
                    selected_cols.append(col)
                    
            if not selected_cols:
                messagebox.showwarning("警告", "未在数据中匹配到选中的列！")
                return
                
            save_df = df_to_plot[selected_cols]
            
            # 计算保存文件的完整路径，默认保存在当前打开文件的目录下
            opened_file = self.file_path.get()
            if opened_file and os.path.exists(opened_file):
                save_dir = os.path.dirname(os.path.abspath(opened_file))
                file_name = os.path.join(save_dir, "Plot_Data.xlsx")
            else:
                file_name = "Plot_Data.xlsx"
                
            next_sheet = "sheet1"
            
            if os.path.exists(file_name):
                try:
                    wb = openpyxl.load_workbook(file_name, read_only=True, keep_links=False)
                    sheet_names = wb.sheetnames
                    wb.close()
                    
                    import re
                    max_num = 0
                    for name in sheet_names:
                        match = re.match(r'^sheet(\d+)$', name, re.IGNORECASE)
                        if match:
                            num = int(match.group(1))
                            if num > max_num:
                                max_num = num
                    next_sheet = f"sheet{max_num + 1}"
                except Exception as e:
                    self.logger.error(f"读取已有 Excel 的 sheet 结构失败: {str(e)}")
            
            # 写入 Excel 文件
            if os.path.exists(file_name):
                with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    save_df.to_excel(writer, sheet_name=next_sheet, index=False)
            else:
                with pd.ExcelWriter(file_name, mode='w', engine='openpyxl') as writer:
                    save_df.to_excel(writer, sheet_name=next_sheet, index=False)
                    
            messagebox.showinfo("成功", f"绘图数据已保存至 {os.path.abspath(file_name)} 中的 {next_sheet}！")
            self.logger.info(f"保存绘图数据成功: {file_name} -> {next_sheet}")
            
        except Exception as e:
            self.logger.error(f"保存绘图数据时发生异常: {str(e)}")
            messagebox.showerror("错误", f"保存数据失败: {str(e)}")

    def save_settings(self):
        """保存当前设置"""
        settings = {
            'font_family': self.font_family.get(),
            'font_size': self.font_size.get(),
            'legend_y': self.legend_y.get(),
            'legend_x_positions': self.legend_x_positions_str.get(),
            'frame_width': self.frame_width.get(),
            'line_width': self.line_width.get(),
            'auto_downsample': self.auto_downsample.get(),
            'max_plot_points': self.max_plot_points.get()
        }
        try:
            with open('settings.json', 'w') as f:
                json.dump(settings, f)
        except Exception as e:
            self.logger.error(f"保存设置失败: {str(e)}")

    def load_settings(self):
        """加载保存的设置"""
        self._is_loading_settings = True
        try:
            with open('settings.json', 'r') as f:
                settings = json.load(f)
                self.font_family.set(settings.get('font_family', 'SimHei'))
                self.font_size.set(settings.get('font_size', '16'))
                self.legend_y.set(settings.get('legend_y', '1.02'))
                self.legend_x_positions_str.set(settings.get('legend_x_positions', "0, 0.3, 0.6"))
                self.frame_width.set(settings.get('frame_width', '1.5'))
                self.line_width.set(settings.get('line_width', '1.5'))
                self.auto_downsample.set(settings.get('auto_downsample', True))
                self.max_plot_points.set(settings.get('max_plot_points', '10000'))
        except FileNotFoundError:
            pass  # 使用默认设置
        except Exception as e:
            self.logger.error(f"加载设置失败: {str(e)}")
        finally:
            self._is_loading_settings = False
            # 在设置加载完成后，手动进行一次总的字体和图例更新，避免多次重绘
            self.update_font_and_plot()

    def toggle_legend(self):
        """切换图例的显示状态"""
        try:
            if hasattr(self, 'ax'):
                # 获取所有轴对象
                axes = [self.ax]
                for axis in self.ax.figure.axes:
                    if axis != self.ax:
                        axes.append(axis)
                
                # 切换图例显示状态
                if self.legend_visible.get():
                    # 重新创建图例
                    self.update_legend_only()
                else:
                    # 移除所有图例
                    for ax in axes:
                        if ax.get_legend() is not None:
                            ax.get_legend().remove()
                
                # 重新绘制画布
                self.canvas.draw()
        except Exception as e:
            self.logger.error(f"切换图例显示状态失败: {str(e)}")

    def update_legend_positions(self):
        """更新图例位置（带防抖）"""
        if getattr(self, '_is_loading_settings', False):
            return
        if hasattr(self, '_legend_pos_timer'):
            try:
                self.root.after_cancel(self._legend_pos_timer)
            except Exception:
                pass
                
        def do_update():
            try:
                # 解析输入的字符串
                positions = [float(x.strip()) for x in self.legend_x_positions_str.get().split(',')]
                if len(positions) >= 3:  # 确保有足够的值
                    self.update_legend_only()
            except Exception:
                pass  # 如果解析失败，忽略错误

        self._legend_pos_timer = self.root.after(300, do_update)

    def get_legend_positions(self):
        """获取图例位置列表"""
        try:
            positions = [float(x.strip()) for x in self.legend_x_positions_str.get().split(',')]
            return positions[:3]  # 只返回前三个值
        except Exception:
            return [0, 0.3, 0.6]  # 返回默认值

    def update_status(self, message, clear=False):
        """更新状态信息"""
        if clear:
            self.status_text.delete(1.0, tk.END)
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)  # 滚动到最新信息
        self.root.update()

    def update_listboxes(self):
        """更新所有列表框的内容"""
        try:
            # 删除空列和空格列
            self.result_df = self.result_df.dropna(axis=1, how='all')
            cols_to_drop = []
            for col in self.result_df.columns:
                if self.result_df[col].dtype == 'object':
                    non_nulls = self.result_df[col].dropna()
                    if len(non_nulls) > 0:
                        # 采样前 1000 行快速检测，如果样本已有非空内容则直接保留，避免大数据量下全量解析卡死
                        sample = non_nulls.head(1000).astype(str).str.strip()
                        if (sample != '').any():
                            continue
                        if non_nulls.astype(str).str.strip().eq('').all():
                            cols_to_drop.append(col)
            if cols_to_drop:
                self.result_df = self.result_df.drop(columns=cols_to_drop)
            
            # 添加索引列
            self.result_df['Index'] = self.result_df.index
            
            # 更新列选择下拉框和列表框
            columns = ['Index'] + [col for col in self.result_df.columns if col != 'Index']
            self.x_combo['values'] = columns
            
            # 电池数据默认选择时间差列或时间列作为X轴
            if self.file_type.get() == "battery" and hasattr(self, 'time_col') and self.time_col.get():
                time_col_name = self.time_col.get()
                time_diff_col = f"{time_col_name}_时间差(s)"
                if time_diff_col in self.result_df.columns:
                    self.x_combo.set(time_diff_col)
                elif time_col_name in self.result_df.columns:
                    self.x_combo.set(time_col_name)
                else:
                    self.x_combo.set(columns[0])
            else:
                self.x_combo.set(columns[0])  # 默认选择索引列
            
            # 清除所有Listbox
            for listbox in self.y_listboxes:
                listbox.delete(0, tk.END)
                # 添加除X轴和索引列外的所有列作为Y轴选项
                for col in self.result_df.columns:
                    if col != 'Index':  # 不将索引列添加到Y轴选项中
                        listbox.insert(tk.END, col)
                    
        except Exception as e:
            self.update_status(f"更新列表失败: {str(e)}")

    def on_closing(self):
        """窗口关闭时的清理工作"""
        try:
            # 清理资源
            self.clear_memory()
            # 保存设置
            self.save_settings()
            # 关闭日志
            logging.shutdown()
            # 销毁窗口
            self.root.destroy()
            # 确保程序完全退出
            os._exit(0)
        except Exception as e:
            print(f"关闭程序时出错: {str(e)}")
            os._exit(1)

    def update_y_axis(self, axis):
        """更新指定Y轴的范围和标题（带防抖）"""
        if getattr(self, '_is_loading_settings', False):
            return
            
        if not hasattr(self, '_y_axis_timers'):
            self._y_axis_timers = {}
            
        if axis in self._y_axis_timers:
            try:
                self.root.after_cancel(self._y_axis_timers[axis])
            except Exception:
                pass
                
        def do_update():
            try:
                if not hasattr(self, 'fig') or not self.fig.axes:
                    return
                
                # 获取对应的轴对象
                axes = self.fig.axes
                target_ax = None
                
                if axis == 0:  # Y1轴
                    target_ax = self.ax
                elif axis == 1:  # Y2轴
                    # 获取Y2轴 - 找到第一个右侧轴
                    for ax in axes:
                        if (ax != self.ax and 
                            ax.spines['right'].get_position()[0] == 'outward' and 
                            ax.spines['right'].get_position()[1] == 0):
                            target_ax = ax
                            break
                elif axis == 2:  # Y3轴
                    # 获取Y3轴 - 找到第二个右侧轴（位置在70）
                    for ax in axes:
                        if (ax != self.ax and 
                            ax.spines['right'].get_position()[0] == 'outward' and 
                            ax.spines['right'].get_position()[1] == 70):
                            target_ax = ax
                            break
                
                if target_ax:
                    # 只更新范围和标题，不清除或重绘
                    try:
                        ymin = float(self.y_settings[axis]['min'].get())
                        ymax = float(self.y_settings[axis]['max'].get())
                        if ymin < ymax:
                            # 保存当前的数据线
                            lines = target_ax.get_lines()
                            # 更新范围
                            target_ax.set_ylim(ymin, ymax)
                            # 确保数据线保持可见
                            for line in lines:
                                line.set_visible(True)
                            
                            # 更新标题，使用自动 labelpad 避免与刻度重合
                            pad_val = 10 if axis == 0 else 15
                            target_ax.set_ylabel(self.y_settings[axis]['title'].get(),
                                               fontsize=int(self.font_size.get()),
                                               fontfamily=self.font_family.get(),
                                               color='black',
                                               labelpad=pad_val)
                            
                            # 重新绘制画布
                            self.canvas.draw()
                    except (ValueError, TypeError):
                        pass
                
            except Exception as e:
                print(f"更新Y{axis+1}轴失败: {str(e)}")

        self._y_axis_timers[axis] = self.root.after(300, do_update)

    def convert_csv_to_xlsx(self):
        """将选中的 CSV 文件转换为 XLSX 文件"""
        file_path = self.file_path.get()
        if not file_path:
            messagebox.showwarning("警告", "请先选择需要转换的 CSV 文件！")
            return
            
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext != '.csv':
            messagebox.showwarning("警告", "当前选择的文件不是 CSV 格式！")
            return
            
        self.set_buttons_state(False)
        threading.Thread(target=self._bg_convert_csv_to_xlsx, args=(file_path,), daemon=True).start()

    def _bg_convert_csv_to_xlsx(self, file_path):
        import polars as pl
        import xlsxwriter
        
        filename = os.path.basename(file_path)
        self.msg_queue.put({'type': 'status', 'message': f"开始转换 CSV: {filename}...", 'clear': True})
        
        try:
            start_time = time.time()
            excel_file = os.path.splitext(file_path)[0] + ".xlsx"
            
            MAX_EXCEL_ROWS = 1048575
            workbook = xlsxwriter.Workbook(excel_file, {'constant_memory': True, 'strings_to_numbers': True})
            
            # 使用 Polars 批量读取 CSV 以提高效率
            reader = pl.read_csv_batched(file_path, batch_size=MAX_EXCEL_ROWS, ignore_errors=True, infer_schema_length=10000)
            sheet_idx = 1
            total_rows_written = 0
            current_sheet_rows = 0
            worksheet = None
            headers = None
            
            while True:
                batches = reader.next_batches(1)
                if not batches:
                    break
                df = batches[0]
                
                if headers is None:
                    headers = df.columns
                    worksheet = workbook.add_worksheet(f"Sheet{sheet_idx}")
                    worksheet.write_row(0, 0, headers)
                
                for row in df.iter_rows():
                    if current_sheet_rows >= MAX_EXCEL_ROWS:
                        sheet_idx += 1
                        worksheet = workbook.add_worksheet(f"Sheet{sheet_idx}")
                        worksheet.write_row(0, 0, headers)
                        current_sheet_rows = 0
                        self.msg_queue.put({'type': 'status', 'message': f"当前工作表已满，已创建 Sheet{sheet_idx}..."})
                        
                    worksheet.write_row(current_sheet_rows + 1, 0, row)
                    current_sheet_rows += 1
                    total_rows_written += 1
                    
                self.msg_queue.put({'type': 'status', 'message': f"已处理并写入数据：共 {total_rows_written:,} 行"})
                
            if worksheet is None:
                worksheet = workbook.add_worksheet("Sheet1")
                
            self.msg_queue.put({'type': 'status', 'message': f"正在压缩并保存 XLSX 文件..."})
            workbook.close()
            
            elapsed = time.time() - start_time
            msg = f"转换完成！已保存为：{os.path.basename(excel_file)}\n共写入 {total_rows_written:,} 行，耗时: {elapsed:.2f}秒"
            self.msg_queue.put({'type': 'status', 'message': msg})
            
            # 主线程弹窗提示成功
            self.root.after(0, lambda: messagebox.showinfo("成功", f"CSV 转换 Excel 成功！\n保存至：{excel_file}\n耗时: {elapsed:.2f}秒"))
            
            # 自动把界面上的文件路径改为生成好的 .xlsx，方便直接读取
            self.root.after(0, lambda: self.file_path.set(excel_file))
            self.root.after(0, self.update_file_type)
            
        except Exception as e:
            error_msg = f"CSV 转换失败: {str(e)}"
            self.logger.error(error_msg)
            self.msg_queue.put({'type': 'error', 'message': error_msg})
            self.root.after(0, lambda: messagebox.showerror("错误", f"CSV 转换失败：{str(e)}"))
        finally:
            self.root.after(0, lambda: self.set_buttons_state(True))

def main():
    # 启用高DPI适配以支持高分辨率屏幕缩放
    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            import ctypes
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

    root = tk.Tk()
    app = PlotterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
