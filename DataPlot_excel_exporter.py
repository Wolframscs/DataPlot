import os
import openpyxl
import pandas as pd
import numpy as np
from tkinter import messagebox

class ExcelExporterMixin:
    def save_cycle_compare_data(self):
        """保存循环对比的数据到 Excel"""
        try:
            cycle_col_name = self.cycle_col.get()
            step_col_name = self.step_col.get()
            time_col_name = self.time_col.get()
            voltage_col_name = self.voltage_col.get()
            current_col_name = self.current_col.get()

            # Ensure columns exist
            for name, col in [("循环列", cycle_col_name), ("工步列", step_col_name), 
                              ("时间列", time_col_name), ("电压列", voltage_col_name), 
                              ("电流列", current_col_name)]:
                if not col or col not in self.result_df.columns:
                    messagebox.showwarning("警告", f"未找到{name} '{col}'，无法计算保存。")
                    return

            try:
                max_cycle = int(pd.to_numeric(self.result_df[cycle_col_name], errors='coerce').max())
            except Exception:
                max_cycle = 1

            cycle_str = self.cycle_compare_range_var.get()
            cycles = self.parse_cycles(cycle_str, max_cycle)
            if not cycles:
                messagebox.showwarning("警告", "未指定有效的循环号范围。")
                return

            step_val = self.step_filter.get()

            try:
                time_step = float(self.time_step_var.get())
            except ValueError:
                time_step = 10.0

            filter_type = self.filter_type_var.get()
            try:
                w_size = int(self.filter_window_var.get())
            except ValueError:
                w_size = 15
            if w_size % 2 == 0: w_size += 1
            try:
                p_order = int(self.sg_poly_var.get())
            except ValueError:
                p_order = 2
            if p_order >= w_size: p_order = w_size - 1

            has_scipy = True
            try:
                from scipy.signal import savgol_filter
            except ImportError:
                has_scipy = False

            # Collect results
            save_dfs = []
            for c in cycles:
                df_c = self.result_df[self.result_df[cycle_col_name] == c].copy()
                # Apply scale factors if specified
                v_scale = self.safe_float_convert(self.voltage_scale_var.get(), 1.0)
                c_scale = self.safe_float_convert(self.current_scale_var.get(), 1.0)
                if v_scale != 1.0 and voltage_col_name in df_c.columns:
                    df_c[voltage_col_name] = pd.to_numeric(df_c[voltage_col_name], errors='coerce') * v_scale
                if c_scale != 1.0 and current_col_name in df_c.columns:
                    df_c[current_col_name] = pd.to_numeric(df_c[current_col_name], errors='coerce') * c_scale
                if step_val != "全部" and step_val != "":
                    df_c = self.compute_step_time(df_c, cycle_col_name, step_col_name, time_col_name)
                    df_c = df_c[df_c[step_col_name].astype(str) == str(step_val)]

                if df_c.empty:
                    continue

                df_c = df_c.sort_values(by=time_col_name)

                time_diff_col = f"{time_col_name}_时间差(s)"
                if time_diff_col in df_c.columns:
                    t_vals = pd.to_numeric(df_c[time_diff_col], errors='coerce').values
                else:
                    t_vals = pd.to_numeric(df_c[time_col_name], errors='coerce').values
                
                valid_t = ~np.isnan(t_vals)
                if not np.any(valid_t):
                    continue
                t_vals = t_vals[valid_t]
                df_c = df_c.iloc[valid_t]

                t_rel = t_vals - t_vals[0]

                # Get CC polarity settings
                cc_polarity = self.cc_polarity_var.get()

                # Check if we should group by step
                use_step_grouping = False
                if step_col_name in df_c.columns:
                    try:
                        unique_steps = df_c[step_col_name].dropna().unique()
                        if len(unique_steps) <= 100:
                            use_step_grouping = True
                    except Exception:
                        pass

                # Get current multiplier for this cycle to align CC polarity
                multiplier = self.get_current_multiplier(df_c, current_col_name, voltage_col_name, step_col_name, cc_polarity)

                # Capacity integration
                cap_vals = np.zeros(len(df_c))
                dt = np.diff(t_rel, prepend=0)
                curr_vals = pd.to_numeric(df_c[current_col_name], errors='coerce').fillna(0).values

                if use_step_grouping:
                    df_c_indices = np.arange(len(df_c))
                    for s in unique_steps:
                        mask = df_c[step_col_name] == s
                        if not np.any(mask):
                            continue
                        step_curr = curr_vals[mask]
                        step_dt = dt[mask]
                        
                        step_dq = (step_curr * multiplier * step_dt) / 3600.0
                        step_cap = np.cumsum(step_dq)
                        cap_vals[mask] = step_cap
                else:
                    dq = (curr_vals * multiplier * dt) / 3600.0
                    cap_vals = np.cumsum(dq)

                # 只计算电流不为0的数据
                curr_vals_raw = pd.to_numeric(df_c[current_col_name], errors='coerce').fillna(0).values
                non_zero_mask = curr_vals_raw != 0
                if not np.any(non_zero_mask):
                    continue
                
                df_c = df_c[non_zero_mask]
                cap_vals = cap_vals[non_zero_mask]
                t_rel = t_rel[non_zero_mask]
                u_vals = pd.to_numeric(df_c[voltage_col_name], errors='coerce').values

                # Calculate dQ/dV and dV/dQ
                target_t = t_rel + time_step
                j_indices = np.searchsorted(t_rel, target_t)
                j_indices = np.minimum(j_indices, len(t_rel) - 1)
                valid = target_t <= t_rel[-1]

                if not np.any(valid):
                    dq_dv_smooth = np.zeros_like(t_rel)
                    dv_dq_smooth = np.zeros_like(t_rel)
                else:
                    dq_diff = cap_vals[j_indices] - cap_vals
                    du_diff = u_vals[j_indices] - u_vals

                    dq_dv = np.where(du_diff != 0, dq_diff / du_diff, np.nan)
                    dv_dq = np.where(dq_diff != 0, du_diff / dq_diff, np.nan)

                    # Smooth
                    s_dq_dv = pd.Series(dq_dv).interpolate(limit_direction='both').fillna(0).values
                    s_dv_dq = pd.Series(dv_dq).interpolate(limit_direction='both').fillna(0).values

                    if filter_type == "Savitzky-Golay" and has_scipy:
                        local_w_size = w_size
                        local_p_order = p_order
                        if len(s_dq_dv) < local_w_size:
                            local_w_size = len(s_dq_dv)
                            if local_w_size % 2 == 0:
                                local_w_size = max(1, local_w_size - 1)
                            if local_p_order >= local_w_size:
                                local_p_order = max(0, local_w_size - 1)
                        
                        if local_w_size > local_p_order and local_w_size >= 3:
                            dq_dv_smooth = savgol_filter(s_dq_dv, window_length=local_w_size, polyorder=local_p_order)
                            dv_dq_smooth = savgol_filter(s_dv_dq, window_length=local_w_size, polyorder=local_p_order)
                        else:
                            dq_dv_smooth = pd.Series(s_dq_dv).rolling(window=min(3, len(s_dq_dv)), center=True, min_periods=1).mean().values
                            dv_dq_smooth = pd.Series(s_dv_dq).rolling(window=min(3, len(s_dv_dq)), center=True, min_periods=1).mean().values
                    elif filter_type == "移动平均":
                        dq_dv_smooth = pd.Series(s_dq_dv).rolling(window=w_size, center=True, min_periods=1).mean().values
                        dv_dq_smooth = pd.Series(s_dv_dq).rolling(window=w_size, center=True, min_periods=1).mean().values
                    elif filter_type == "中值滤波":
                        dq_dv_smooth = pd.Series(s_dq_dv).rolling(window=w_size, center=True, min_periods=1).median().values
                        dv_dq_smooth = pd.Series(s_dv_dq).rolling(window=w_size, center=True, min_periods=1).median().values
                    else:
                        dq_dv_smooth = s_dq_dv
                        dv_dq_smooth = s_dv_dq

                # Create output df for this cycle
                data_dict = {
                    '循环号': c,
                    '工步': df_c[step_col_name].values,
                    '时间/s': t_rel,
                }
                if '工步时间' in df_c.columns:
                    data_dict['工步时间/s'] = df_c['工步时间'].values
                data_dict.update({
                    '累计容量/Ah': cap_vals,
                    '电压/V': u_vals,
                    '电流/A': pd.to_numeric(df_c[current_col_name], errors='coerce').values,
                    'dQ_dV/(Ah_V)': dq_dv_smooth,
                    'dV_dQ/(V_Ah)': dv_dq_smooth
                })
                c_df = pd.DataFrame(data_dict)
                save_dfs.append(c_df)

            if not save_dfs:
                messagebox.showwarning("警告", "没有可保存的计算数据。")
                return

            save_dfs_reset = [df.reset_index(drop=True) for df in save_dfs]
            final_df = pd.concat(save_dfs_reset, axis=1)

            opened_file = self.file_path.get()
            base_name = "BATTERY_Cycle_Compare_Data.xlsx"
            if opened_file and os.path.exists(opened_file):
                save_dir = os.path.dirname(os.path.abspath(opened_file))
                file_name = os.path.join(save_dir, base_name)
            else:
                file_name = base_name

            next_sheet = "sheet1"
            if os.path.exists(file_name):
                try:
                    wb = openpyxl.load_workbook(file_name, read_only=True, keep_links=False)
                    sheet_names = wb.sheetnames
                    wb.close()
                    import re
                    max_num = 0
                    for name in sheet_names:
                        match = re.match(r'^sheet(\d+)$', name, re.IGNORECASE)
                        if match:
                            num = int(match.group(1))
                            if num > max_num: max_num = num
                    next_sheet = f"sheet{max_num + 1}"
                except Exception:
                    pass

            if os.path.exists(file_name):
                with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    final_df.to_excel(writer, sheet_name=next_sheet, index=False)
            else:
                with pd.ExcelWriter(file_name, mode='w', engine='openpyxl') as writer:
                    final_df.to_excel(writer, sheet_name=next_sheet, index=False)

            messagebox.showinfo("成功", f"循环对比数据已保存至 {os.path.abspath(file_name)} 中的 {next_sheet}！")

        except Exception as e:
            messagebox.showerror("错误", f"保存循环对比数据失败: {str(e)}")

    def save_plot_data(self):
        """将当前绘图数据保存为 xlsx 文件"""
        try:
            if self.result_df is None:
                messagebox.showwarning("警告", "当前无有效数据，请先载入并计算数据！")
                return
            
            if self.file_type.get() == "battery" and self.cycle_compare_var.get():
                self.save_cycle_compare_data()
                return
            
            x_col = self.x_axis.get()
            if not x_col:
                messagebox.showwarning("警告", "请先选择 X 轴数据！")
                return
                
            y1_cols = self.y_selections[0]
            y2_cols = self.y_selections[1]
            y3_cols = self.y_selections[2]
            all_y_cols = y1_cols + y2_cols + y3_cols
            
            if not all_y_cols:
                messagebox.showwarning("警告", "请先选择至少一个 Y 轴进行绘图！")
                return
                
            df_to_plot = self.result_df
            
            if self.file_type.get() == "battery":
                cycle_col_name = self.cycle_col.get()
                step_col_name = self.step_col.get()
                cycle_val = self.cycle_filter.get()
                step_val = self.step_filter.get()
                
                if cycle_col_name and cycle_col_name in df_to_plot.columns:
                    if cycle_val != "全部" and cycle_val != "":
                        try:
                            try:
                                target_val = int(cycle_val)
                                df_to_plot = df_to_plot[df_to_plot[cycle_col_name] == target_val]
                            except ValueError:
                                try:
                                    target_val = float(cycle_val)
                                    df_to_plot = df_to_plot[df_to_plot[cycle_col_name] == target_val]
                                except ValueError:
                                    df_to_plot = df_to_plot[df_to_plot[cycle_col_name].astype(str) == str(cycle_val)]
                        except Exception:
                            pass
                if step_col_name and step_col_name in df_to_plot.columns:
                    if step_val != "全部" and step_val != "":
                        df_to_plot = self.compute_step_time(df_to_plot, cycle_col_name, step_col_name, self.time_col.get())
                        try:
                            df_to_plot = df_to_plot[df_to_plot[step_col_name] == step_val]
                        except Exception:
                            try:
                                df_to_plot = df_to_plot[df_to_plot[step_col_name].astype(str) == str(step_val)]
                            except Exception:
                                pass
            
            if df_to_plot.empty:
                messagebox.showwarning("警告", "筛选后的绘图数据为空，无法保存！")
                return
                
            selected_cols = []
            if x_col and x_col in df_to_plot.columns:
                selected_cols.append(x_col)
            for col in all_y_cols:
                if col and col in df_to_plot.columns and col not in selected_cols:
                    selected_cols.append(col)
                    
            if not selected_cols:
                messagebox.showwarning("警告", "未在数据中匹配到选中的列！")
                return
                
            save_df = df_to_plot[selected_cols].copy()
            save_df.columns = [self.clean_legend_label(col) for col in save_df.columns]
            
            opened_file = self.file_path.get()
            panel_prefix = "FLOEFD"
            if self.file_type.get() == "processed":
                panel_prefix = "GENERAL"
            elif self.file_type.get() == "battery":
                panel_prefix = "BATTERY"
                
            base_name = f"{panel_prefix}_Plot_Data.xlsx"
            if opened_file and os.path.exists(opened_file):
                save_dir = os.path.dirname(os.path.abspath(opened_file))
                file_name = os.path.join(save_dir, base_name)
            else:
                file_name = base_name
                
            next_sheet = "sheet1"
            
            if os.path.exists(file_name):
                try:
                    wb = openpyxl.load_workbook(file_name, read_only=True, keep_links=False)
                    sheet_names = wb.sheetnames
                    wb.close()
                    
                    import re
                    max_num = 0
                    for name in sheet_names:
                        match = re.match(r'^sheet(\d+)$', name, re.IGNORECASE)
                        if match:
                            num = int(match.group(1))
                            if num > max_num:
                                max_num = num
                    next_sheet = f"sheet{max_num + 1}"
                except Exception as e:
                    self.logger.error(f"读取已有 Excel 的 sheet 结构失败: {str(e)}")
            
            if os.path.exists(file_name):
                with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    save_df.to_excel(writer, sheet_name=next_sheet, index=False)
            else:
                with pd.ExcelWriter(file_name, mode='w', engine='openpyxl') as writer:
                    save_df.to_excel(writer, sheet_name=next_sheet, index=False)
                    
            messagebox.showinfo("成功", f"绘图数据已保存至 {os.path.abspath(file_name)} 中的 {next_sheet}！")
            self.logger.info(f"保存绘图数据成功: {file_name} -> {next_sheet}")
            
        except Exception as e:
            self.logger.error(f"保存绘图数据时发生异常: {str(e)}")
            messagebox.showerror("错误", f"保存数据失败: {str(e)}")
